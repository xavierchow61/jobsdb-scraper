"""Telegram bot listener: trigger scrapes and query master DB by chat command.

Reads bot_token + chat_id + master path from config.json (saved by GUI),
or pass them as --token / --chat-id / --master flags.

Run:  python bot_listener.py
"""

from __future__ import annotations

import argparse
import json
import sys
import threading
import time
import traceback
import urllib.error
import urllib.parse
import urllib.request
from collections import Counter
from datetime import datetime
from pathlib import Path

import scraper

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except (AttributeError, OSError):
    pass

APP_DIR = Path(__file__).resolve().parent
CONFIG_PATH = APP_DIR / "config.json"
STATE_PATH = APP_DIR / "bot_state.json"
TELEGRAM_API = "https://api.telegram.org"


def load_state():
    if not STATE_PATH.exists():
        return {}
    try:
        return json.loads(STATE_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {}


def save_state(state):
    try:
        STATE_PATH.write_text(
            json.dumps(state, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    except Exception as e:
        print(f"  save_state error: {e}")

INLINE_SOURCES = [[
    {"text": "JobsDB", "callback_data": "src:jobsdb"},
    {"text": "CTgoodjobs", "callback_data": "src:ctgoodjobs"},
    {"text": "cpjobs", "callback_data": "src:cpjobs"},
]]

INLINE_CP_LOCATIONS = [
    [{"text": "🏙 Hong Kong Island", "callback_data": "loc:Hong Kong Island"}],
    [{"text": "🌃 Kowloon", "callback_data": "loc:Kowloon"}],
    [{"text": "🌄 New Territories", "callback_data": "loc:New Territories"}],
    [{"text": "🏝 Outlying Islands", "callback_data": "loc:Outlying Islands"}],
    [{"text": "⏭ 全港 (跳過)", "callback_data": "loc:__skip__"}],
]

INLINE_CONFIRM = [[
    {"text": "✅ 開始", "callback_data": "confirm:yes"},
    {"text": "❌ 取消", "callback_data": "confirm:no"},
]]

# Curated HK industries. Each entry maps the display label to a per-source keyword
# (whatever each site recognises best). 'other' = ask free-text keyword.
INDUSTRIES = [
    ("會計 Accounting", {"jobsdb": "Accountant", "ctgoodjobs": "Accounting",
                         "cpjobs": "accountant"}),
    ("銀行 Banking", {"jobsdb": "Banking", "ctgoodjobs": "banking-finance",
                       "cpjobs": "banking"}),
    ("金融 Finance", {"jobsdb": "Finance", "ctgoodjobs": "banking-finance",
                      "cpjobs": "finance"}),
    ("審計 Audit", {"jobsdb": "Audit", "ctgoodjobs": "accounting",
                    "cpjobs": "audit"}),
    ("稅務 Tax", {"jobsdb": "Tax", "ctgoodjobs": "accounting",
                  "cpjobs": "tax"}),
    ("IT", {"jobsdb": "IT", "ctgoodjobs": "information-technology",
            "cpjobs": "IT"}),
    ("市場 Marketing", {"jobsdb": "Marketing",
                         "ctgoodjobs": "marketing-public-relations",
                         "cpjobs": "marketing"}),
    ("銷售 Sales", {"jobsdb": "Sales", "ctgoodjobs": "sales",
                    "cpjobs": "sales"}),
    ("人事 HR", {"jobsdb": "Human Resources",
                  "ctgoodjobs": "human-resources",
                  "cpjobs": "human resources"}),
    ("行政 Admin", {"jobsdb": "Administration",
                     "ctgoodjobs": "administration",
                     "cpjobs": "administration"}),
    ("工程 Engineering", {"jobsdb": "Engineer", "ctgoodjobs": "engineering",
                          "cpjobs": "engineer"}),
    ("教育 Education", {"jobsdb": "Teacher", "ctgoodjobs": "education",
                        "cpjobs": "teacher"}),
    ("零售 Retail", {"jobsdb": "Retail", "ctgoodjobs": "retail-merchandise",
                     "cpjobs": "retail"}),
    ("酒店 Hospitality", {"jobsdb": "Hotel",
                           "ctgoodjobs": "hospitality-tourism",
                           "cpjobs": "hospitality"}),
    ("物流 Logistics", {"jobsdb": "Logistics",
                         "ctgoodjobs": "logistics-transportation",
                         "cpjobs": "logistics"}),
    ("法律 Legal", {"jobsdb": "Legal", "ctgoodjobs": "legal",
                    "cpjobs": "legal"}),
]

# Inline keyboard, 2 per row, with a final "其他" row
def _build_industry_keyboard():
    rows = []
    for i in range(0, len(INDUSTRIES), 2):
        row = []
        for label, _ in INDUSTRIES[i:i + 2]:
            row.append({"text": label, "callback_data": f"ind:{label}"})
        rows.append(row)
    rows.append([{"text": "🔤 其他 (自己打 keyword)",
                  "callback_data": "ind:__other__"}])
    return rows


INLINE_INDUSTRIES = _build_industry_keyboard()

INDUSTRY_BY_LABEL = {label: mapping for label, mapping in INDUSTRIES}

BOT_COMMANDS = [
    {"command": "scrape", "description": "搜尋新工作 (一步一步引導)"},
    {"command": "cv", "description": "睇/改 CV 關鍵字 + 年資"},
    {"command": "find", "description": "喺資料庫搵 job (例: /find SAP)"},
    {"command": "top", "description": "配對度最高嘅 N 個 (預設 5)"},
    {"command": "last", "description": "最近爬到嘅 N 個 (預設 5)"},
    {"command": "stats", "description": "資料庫總數 + 來源分佈"},
    {"command": "status", "description": "睇而家有冇爬緊"},
    {"command": "stop", "description": "停止當前爬蟲"},
    {"command": "cancel", "description": "取消 Q&A 流程"},
    {"command": "help", "description": "顯示完整指令列表"},
]


def api_call(token, method, payload=None, timeout=35):
    url = f"{TELEGRAM_API}/bot{token}/{method}"
    data = None
    headers = {}
    if payload is not None:
        data = json.dumps(payload).encode("utf-8")
        headers["Content-Type"] = "application/json"
    req = urllib.request.Request(
        url, data=data, method="POST" if data else "GET", headers=headers,
    )
    try:
        with urllib.request.urlopen(req, timeout=timeout) as r:
            return json.loads(r.read())
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", errors="replace")[:500]
        return {"ok": False, "error_code": e.code, "description": body}
    except Exception as e:
        return {"ok": False, "description": str(e)}


def send_msg(token, chat_id, text, reply_markup=None):
    payload = {
        "chat_id": str(chat_id),
        "text": text,
        "parse_mode": "HTML",
        "disable_web_page_preview": True,
    }
    if reply_markup is not None:
        payload["reply_markup"] = reply_markup
    res = api_call(token, "sendMessage", payload, timeout=20)
    if not res.get("ok"):
        print(
            f"[{datetime.now():%H:%M:%S}] sendMessage FAILED: "
            f"{res.get('description', res)} | text[:60]={text[:60]!r}"
        )
    return res


def answer_callback(token, cb_id, text=None):
    payload = {"callback_query_id": cb_id}
    if text:
        payload["text"] = text
    return api_call(token, "answerCallbackQuery", payload, timeout=10)


def set_bot_commands(token):
    return api_call(token, "setMyCommands", {"commands": BOT_COMMANDS}, timeout=10)


class BotState:
    """Per-user conversation state."""

    def __init__(self):
        self.flow = None
        self.step = None
        self.data = {}

    def reset(self):
        self.flow = None
        self.step = None
        self.data = {}


class Bot:
    def __init__(self, token, chat_id, master_path, cv_path=""):
        self.token = token
        self.chat_id = str(chat_id)
        self.master_path = master_path
        self.cv_path = cv_path
        self.offset = 0
        self.state = BotState()
        self.scrape_thread: threading.Thread | None = None
        self.scrape_stop = threading.Event()
        self.scrape_args = None
        # Persisted last-used scrape choices (source / location / etc.)
        self.persist = load_state()
        self.last = self.persist.setdefault("last", {})

    # ---------- helpers ----------

    def reply(self, text, reply_markup=None):
        return send_msg(self.token, self.chat_id, text, reply_markup=reply_markup)

    def reply_buttons(self, text, buttons):
        return self.reply(text, reply_markup={"inline_keyboard": buttons})

    def _open_master(self):
        if load_workbook is None:
            self.reply("❌ openpyxl 未裝。")
            return None
        path = Path(self.master_path)
        if not path.exists():
            self.reply(f"❌ Master 唔存在: {path.name}")
            return None
        try:
            return load_workbook(path, read_only=True, data_only=True)
        except Exception as e:
            self.reply(f"❌ 開唔到 master: {e}")
            return None

    def _master_rows(self):
        wb = self._open_master()
        if wb is None:
            return None, None
        ws = wb.active
        if ws.max_row is None or ws.max_row < 2:
            self.reply("Master 仲係空白。")
            wb.close()
            return None, None
        headers = [c.value for c in ws[1]]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()
        return headers, rows

    # ---------- run loop ----------

    def run(self):
        print(f"[{datetime.now():%H:%M:%S}] Bot listening for chat_id={self.chat_id}")

        # Sanity check: who am I + is a webhook blocking getUpdates?
        me = api_call(self.token, "getMe", timeout=10)
        if me.get("ok"):
            u = me.get("result", {})
            print(f"  Bot identity: @{u.get('username')} (id={u.get('id')})")
        else:
            print(f"  getMe failed: {me}")

        wh = api_call(self.token, "getWebhookInfo", timeout=10)
        if wh.get("ok"):
            info = wh.get("result", {})
            wh_url = info.get("url") or ""
            if wh_url:
                print(
                    f"  WEBHOOK is set to: {wh_url}\n"
                    "  → Deleting webhook so long-polling works..."
                )
                dr = api_call(self.token, "deleteWebhook",
                              {"drop_pending_updates": False}, timeout=10)
                print(f"  deleteWebhook -> {dr}")
            else:
                print("  No webhook (good — polling will work).")

        set_bot_commands(self.token)
        self.reply(
            "🤖 <b>Bot online</b>\n"
            "打 /help 睇有咩可以做，或者撳輸入框旁邊個 menu icon。"
        )
        backoff = 1
        poll_count = 0
        allowed_updates_json = urllib.parse.quote(
            json.dumps([
                "message", "edited_message", "callback_query",
                "channel_post", "inline_query",
            ])
        )
        while True:
            res = api_call(
                self.token,
                f"getUpdates?offset={self.offset}&timeout=30"
                f"&allowed_updates={allowed_updates_json}",
                timeout=40,
            )
            if not res.get("ok"):
                desc = res.get("description", "?")
                print(f"[{datetime.now():%H:%M:%S}] getUpdates failed: {desc}")
                time.sleep(min(backoff, 30))
                backoff = min(backoff * 2, 60)
                continue
            backoff = 1
            updates = res.get("result", [])
            poll_count += 1
            if updates:
                print(
                    f"[{datetime.now():%H:%M:%S}] poll #{poll_count}: "
                    f"got {len(updates)} update(s)"
                )
            elif poll_count % 6 == 0:
                # Heartbeat every ~3 min so we know polling is alive
                print(f"[{datetime.now():%H:%M:%S}] poll #{poll_count}: idle")
            for upd in updates:
                self.offset = max(self.offset, upd["update_id"] + 1)
                try:
                    self.handle(upd)
                except Exception as e:
                    print(f"handler error: {e}")
                    traceback.print_exc()
                    try:
                        self.reply(f"⚠ Internal error: {e}")
                    except Exception:
                        pass

    def handle(self, upd):
        kinds = [k for k in (
            "message", "edited_message", "callback_query",
            "channel_post", "edited_channel_post", "my_chat_member",
            "chat_member", "inline_query",
        ) if k in upd]
        print(f"[{datetime.now():%H:%M:%S}] update kinds={kinds}")
        msg = upd.get("message") or upd.get("edited_message")
        if msg:
            return self.handle_message(msg)
        cb = upd.get("callback_query")
        if cb:
            return self.handle_callback(cb)
        print(f"  -> unhandled update payload: {json.dumps(upd)[:200]}")

    def handle_message(self, msg):
        chat_id = str(msg.get("chat", {}).get("id"))
        text = (msg.get("text") or "").strip()
        print(
            f"  msg chat={chat_id} (expected {self.chat_id}) "
            f"text={text[:60]!r}"
        )
        if chat_id != self.chat_id:
            print(f"  -> chat_id mismatch, ignoring")
            return
        if not text:
            print(f"  -> empty text, ignoring")
            return

        # Universally available commands
        if text.startswith("/cancel"):
            self.state.reset()
            return self.reply("❌ 已取消")
        if text.startswith("/stop"):
            return self.cmd_stop()
        if text.startswith("/help") or text == "/start":
            return self.cmd_help()
        if text.startswith("/status"):
            return self.cmd_status()
        if text.startswith("/stats"):
            return self.cmd_stats()

        # In-flow text
        if self.state.flow == "scrape":
            return self.scrape_handle_text(text)
        if self.state.flow == "find":
            return self.find_handle_text(text)

        # Idle commands
        if text == "/scrape" or text.startswith("/scrape "):
            return self.cmd_scrape(text)
        if text == "/find" or text.startswith("/find "):
            return self.cmd_find(text)
        if text.startswith("/last"):
            return self.cmd_last(text)
        if text.startswith("/top"):
            return self.cmd_top(text)
        if text == "/cv" or text.startswith("/cv "):
            return self.cmd_cv(text)

        self.reply("❓ 唔識呢個指令。打 /help 睇 menu。")

    def handle_callback(self, cb):
        cb_id = cb["id"]
        chat_id = str(cb.get("message", {}).get("chat", {}).get("id"))
        data = cb.get("data") or ""
        print(
            f"[{datetime.now():%H:%M:%S}] callback "
            f"chat={chat_id} data={data!r} state={self.state.flow}/{self.state.step}"
        )
        if chat_id != self.chat_id:
            return answer_callback(self.token, cb_id, "Unauthorized")
        ack = answer_callback(self.token, cb_id)
        if not ack.get("ok"):
            print(f"  answerCallback failed: {ack}")

        if data.startswith("src:"):
            return self.scrape_set_source(data.split(":", 1)[1])
        if data.startswith("ind:"):
            return self.scrape_set_industry(data.split(":", 1)[1])
        if data.startswith("loc:"):
            loc = data.split(":", 1)[1]
            if loc == "__skip__":
                loc = ""
            return self.scrape_set_location(loc)
        if data.startswith("mp:"):
            try:
                self.state.data["max_pages"] = int(data.split(":", 1)[1])
            except ValueError:
                return
            return self.scrape_ask_threshold()
        if data.startswith("th:"):
            try:
                self.state.data["match_threshold"] = float(data.split(":", 1)[1])
            except ValueError:
                return
            return self.scrape_ask_max_telegram()
        if data.startswith("mt:"):
            try:
                self.state.data["telegram_max"] = int(data.split(":", 1)[1])
            except ValueError:
                return
            return self.scrape_show_confirm()
        if data == "confirm:yes":
            return self.scrape_start_now()
        if data == "confirm:no":
            self.state.reset()
            return self.reply("❌ 已取消")
        if data.startswith("act:"):
            return self.handle_action_callback(cb_id, cb, data)
        if data == "noop":
            return

    # ---------- /cv  -- view + edit CV keywords ----------

    def _load_cv_profile_or_reply(self):
        """Return (profile, save_path) or None after sending an error reply."""
        try:
            import cv_match
        except ImportError:
            self.reply("⚠ cv_match 唔可用 (pdfminer.six 未裝?)")
            return None
        if not self.cv_path:
            self.reply(
                "⚠ 未設 CV 檔。喺 GUI 揀 CV PDF 再撳 Start 一次儲入 config，"
                "然後重啟 listener。"
            )
            return None
        profile = cv_match.load_cv(self.cv_path)
        if profile is None:
            self.reply(f"⚠ 讀唔到 CV: {self.cv_path}")
            return None
        return profile, cv_match.profile_json_path(self.cv_path)

    def cmd_cv(self, text):
        parts = text.split(maxsplit=1)
        sub_and_rest = parts[1].split(maxsplit=1) if len(parts) > 1 else []
        sub = (sub_and_rest[0].lower() if sub_and_rest else "")
        rest = (sub_and_rest[1] if len(sub_and_rest) > 1 else "")

        if sub in ("", "list", "show"):
            return self.cv_list()
        if sub == "add":
            return self.cv_modify(rest, add=True)
        if sub in ("del", "delete", "remove", "rm"):
            return self.cv_modify(rest, add=False)
        if sub == "years":
            return self.cv_set_years(rest)
        if sub in ("reset", "reload"):
            return self.cv_reset()
        return self.reply(
            "用法:\n"
            "<code>/cv</code> — show 而家所有 keywords\n"
            "<code>/cv add hkfrs, treasury, sap b1</code> — 加 (用 , 分開)\n"
            "<code>/cv del mandarin, cantonese</code> — 刪\n"
            "<code>/cv years 8</code> — 改年資\n"
            "<code>/cv reset</code> — 重新由 PDF 抽取 (會丟失 custom)"
        )

    def cv_list(self):
        loaded = self._load_cv_profile_or_reply()
        if loaded is None:
            return
        profile, _ = loaded
        kws = sorted(profile.keywords)
        years_str = f"{profile.years} 年" if profile.years else "?"
        head = (
            f"📄 <b>CV Profile</b>\n"
            f"Source: <code>{Path(self.cv_path).name}</code>\n"
            f"Years: <b>{years_str}</b>\n"
            f"Keywords: <b>{len(kws)}</b>\n\n"
        )
        if not kws:
            return self.reply(head + "<i>(冇 keyword)</i>")
        # Chunk if very long
        body_lines = []
        for kw in kws:
            body_lines.append(f"• {self._esc(kw)}")
        body = "\n".join(body_lines)
        # Telegram message limit ~4096 chars; split if needed
        if len(head) + len(body) > 3800:
            mid = len(body) // 2
            split_at = body.rfind("\n", 0, mid)
            self.reply(head + body[:split_at])
            self.reply(body[split_at + 1:])
        else:
            self.reply(head + body)

    def cv_modify(self, rest, add):
        if not rest.strip():
            return self.reply(
                "用法:\n"
                "<code>/cv add hkfrs, treasury, sap b1</code>\n"
                "<code>/cv del mandarin, excel</code>"
            )
        loaded = self._load_cv_profile_or_reply()
        if loaded is None:
            return
        profile, save_path = loaded
        try:
            import cv_match
        except ImportError:
            return self.reply("⚠ cv_match 唔可用")
        kws_in = {k.strip().lower() for k in rest.split(",") if k.strip()}
        if not kws_in:
            return self.reply("⚠ 冇 valid keyword (用 , 分開)")
        before = len(profile.keywords)
        if add:
            new = kws_in - profile.keywords
            profile.keywords |= kws_in
            change = f"加咗 {len(new)} 個新 keyword"
        else:
            removed = kws_in & profile.keywords
            profile.keywords -= kws_in
            not_found = kws_in - removed
            change = f"刪咗 {len(removed)} 個 keyword"
            if not_found:
                change += f"\n(唔存在嘅: {', '.join(sorted(not_found))})"
        try:
            cv_match.save_profile(profile, save_path)
        except Exception as e:
            return self.reply(f"⚠ 儲存失敗: {e}")
        self.reply(
            f"✅ {change}\n"
            f"Profile 總共而家有 <b>{len(profile.keywords)}</b> keywords "
            f"(之前 {before})"
        )

    def cv_set_years(self, rest):
        loaded = self._load_cv_profile_or_reply()
        if loaded is None:
            return
        profile, save_path = loaded
        try:
            years = int(rest.strip())
        except ValueError:
            return self.reply("用法: <code>/cv years 8</code>")
        try:
            import cv_match
            profile.years = years
            cv_match.save_profile(profile, save_path)
        except Exception as e:
            return self.reply(f"⚠ 儲存失敗: {e}")
        self.reply(f"✅ 年資設為 <b>{years}</b> 年")

    def cv_reset(self):
        if not self.cv_path:
            return self.reply("⚠ 未設 CV 檔。")
        try:
            import cv_match
            profile = cv_match.load_cv(self.cv_path, use_saved_profile=False)
            if profile is None:
                return self.reply("⚠ 讀唔到 CV")
            cv_match.save_profile(profile)
        except Exception as e:
            return self.reply(f"⚠ Reset 失敗: {e}")
        self.reply(
            f"✅ 由 PDF 重新讀取: {len(profile.keywords)} keywords, "
            f"{profile.years} 年"
        )

    # ---------- action callbacks (Save / Hide / Apply) ----------

    def handle_action_callback(self, cb_id, cb, data):
        # data like "act:save:4256107"
        parts = data.split(":", 2)
        if len(parts) != 3:
            return answer_callback(self.token, cb_id, "Bad action")
        _, action, jd = parts
        column = {
            "save": "Saved", "hide": "Hidden", "apply": "Applied",
        }.get(action)
        if not column:
            return answer_callback(self.token, cb_id, "Unknown action")
        ok = scraper.master_set_status(self.master_path, jd, column)
        if not ok:
            return answer_callback(
                self.token, cb_id, "✗ failed (file open or JD missing)"
            )
        new_label = {
            "save": "⭐ Saved", "hide": "🚫 Hidden",
            "apply": "✅ Applied",
        }[action]
        # Update the card's buttons in place
        msg = cb.get("message", {})
        message_id = msg.get("message_id")
        if message_id:
            new_kb = {"inline_keyboard": [[
                {"text": new_label, "callback_data": "noop"},
            ]]}
            api_call(self.token, "editMessageReplyMarkup", {
                "chat_id": self.chat_id,
                "message_id": message_id,
                "reply_markup": new_kb,
            }, timeout=10)
        answer_callback(self.token, cb_id, new_label)

    # ---------- /help ----------

    def cmd_help(self):
        self.reply(
            "<b>📚 指令列表</b>\n"
            "\n"
            "━━━━━ 🚀 <b>搜尋新工作</b> ━━━━━\n"
            "<b>/scrape</b> — 開始搜尋 (一步一步引導)\n"
            "\n"
            "━━━━━ 📝 <b>編輯 CV 關鍵字</b> ━━━━━\n"
            "<b>/cv</b> — 睇而家所有 keywords + 年資\n"
            "<b>/cv add</b>  — 加 keywords "
            "(例: <code>/cv add hkfrs, treasury, sap b1</code>)\n"
            "<b>/cv del</b> 字眼 — 刪 keywords "
            "(例: <code>/cv del mandarin, cantonese</code>)\n"
            "<b>/cv years</b> 數字 — 設年資 (例: <code>/cv years 8</code>)\n"
            "<b>/cv reset</b> — 由 PDF 重新抽 (會丟失 custom)\n"
            "\n"
            "━━━━━ 🔍 <b>查詢資料庫</b> ━━━━━\n"
            "<b>/find</b> 字眼 — 搵 job (例: <code>/find SAP</code>)\n"
            "<b>/top</b> [N] — 配對度最高嘅 N 個 (預設 5)\n"
            "<b>/last</b> [N] — 最近爬到嘅 N 個 (預設 5)\n"
            "<b>/stats</b> — 總數 + 來源分佈\n"
            "\n"
            "━━━━━ ⚙️ <b>系統</b> ━━━━━\n"
            "<b>/status</b> — 睇而家有冇爬緊\n"
            "<b>/stop</b> — 停止當前爬蟲\n"
            "<b>/cancel</b> — 取消 Q&amp;A 流程\n"
            "<b>/help</b> — 顯示呢個列表\n"
            "\n"
            "<i>💡 每張卡片有 [⭐ Save] [🚫 Hide] [✅ Applied] 按掣，\n"
            "撳一撳即時寫入 Master Excel。</i>"
        )

    # ---------- /scrape ----------

    def cmd_scrape(self, text):
        if self.scrape_thread and self.scrape_thread.is_alive():
            return self.reply(
                "⚠️ 已經有 scrape 跑緊。\n打 /status 睇進度，或 /stop 停止。"
            )

        parts = text.split(maxsplit=3)
        if len(parts) >= 3:
            source = parts[1].lower()
            if source not in scraper.SOURCES:
                return self.reply(
                    f"❌ 唔識 source: <code>{source}</code>\n"
                    f"用 jobsdb / ctgoodjobs / cpjobs"
                )
            kw = parts[2]
            loc = parts[3] if len(parts) >= 4 else ""
            self.state.flow = "scrape"
            self.state.data = {"source": source, "keyword": kw, "location": loc}
            return self.scrape_show_confirm()

        self.state.flow = "scrape"
        self.state.step = "source"
        self.state.data = {}
        rows = [list(INLINE_SOURCES[0])]
        if self.last.get("source"):
            rows.append([{
                "text": f"↺ 同上次 ({self.last['source']})",
                "callback_data": f"src:{self.last['source']}",
            }])
        self.reply_buttons("揀網站:", rows)

    def scrape_set_source(self, source):
        if source not in scraper.SOURCES:
            return self.reply(f"❌ 唔識 source: {source}")
        # Resume / restart flow regardless of previous state
        if self.state.flow != "scrape":
            self.state.flow = "scrape"
            self.state.data = {}
        self.state.data["source"] = source
        self.state.step = "industry"
        self.reply_buttons(
            f"✅ Source: <b>{source}</b>\n\n揀行業:",
            INLINE_INDUSTRIES,
        )

    def scrape_set_industry(self, label):
        """User clicked an industry button (or 其他)."""
        if self.state.flow != "scrape" or not self.state.data.get("source"):
            return self.reply(
                "⚠ Flow 中斷咗。打 /scrape 重新開始。"
            )
        if label == "__other__":
            self.state.step = "keyword"
            return self.reply(
                "🔤 打你想搵嘅 keyword (例: <code>CFA</code>, "
                "<code>senior accountant</code>):"
            )
        mapping = INDUSTRY_BY_LABEL.get(label)
        if not mapping:
            return self.reply(f"❌ 唔識 industry: {label}")
        src = self.state.data["source"]
        keyword = mapping.get(src) or label.split()[0]
        self.state.data["keyword"] = keyword
        self.state.data["industry_label"] = label
        # Proceed to location step
        self.state.step = "location"
        src = self.state.data.get("source")
        if src == "cpjobs":
            self.reply_buttons(
                f"✅ Industry: <b>{label}</b> → keyword: <code>{keyword}</code>"
                "\n\n揀地區:", INLINE_CP_LOCATIONS,
            )
        elif src == "ctgoodjobs":
            self.reply(
                f"✅ Industry: <b>{label}</b> → keyword: <code>{keyword}</code>"
                "\n\n打地點 (例: <code>Central</code>, "
                "<code>Tsim Sha Tsui</code>, <code>Tseung Kwan O</code>),\n"
                "或者打 <code>-</code> 跳過 (全港):"
            )
        else:
            self.reply(
                f"✅ Industry: <b>{label}</b> → keyword: <code>{keyword}</code>"
                "\n\n打地點 (例: <code>Central and Western District</code>),\n"
                "或者打 <code>-</code> 跳過 (全港):"
            )

    def scrape_handle_text(self, text):
        if self.state.step == "keyword":
            self.state.data["keyword"] = text
            self.state.step = "location"
            src = self.state.data.get("source")
            if src == "cpjobs":
                self.reply_buttons("揀地區:", INLINE_CP_LOCATIONS)
            elif src == "ctgoodjobs":
                self.reply(
                    "打地點 (例: <code>Central</code>, "
                    "<code>Tsim Sha Tsui</code>, <code>Tseung Kwan O</code>),\n"
                    "或者打 <code>-</code> 跳過 (全港):"
                )
            else:
                self.reply(
                    "打地點 (例: <code>Central and Western District</code>),\n"
                    "或者打 <code>-</code> 跳過 (全港):"
                )
            return
        if self.state.step == "location":
            loc = text.strip()
            if loc == "-":
                loc = ""
            self.state.data["location"] = loc
            return self.scrape_ask_max_pages()
        if self.state.step == "max_pages":
            try:
                self.state.data["max_pages"] = max(0, int(text.strip()))
            except ValueError:
                return self.reply("唔係數字。輸入 1-20 (或 0 = 全部):")
            return self.scrape_ask_threshold()
        if self.state.step == "threshold":
            try:
                self.state.data["match_threshold"] = max(0.0, float(text.strip()))
            except ValueError:
                return self.reply("唔係數字。輸入 0-100:")
            return self.scrape_ask_max_telegram()
        if self.state.step == "max_telegram":
            try:
                self.state.data["telegram_max"] = max(0, int(text.strip()))
            except ValueError:
                return self.reply("唔係數字。輸入 0 (無限) 或正整數:")
            return self.scrape_show_confirm()

    def scrape_set_location(self, loc):
        if self.state.flow != "scrape":
            return self.reply(
                "⚠ Flow 中斷咗（可能 bot 重新啟動過）。打 /scrape 重新開始。"
            )
        self.state.data["location"] = loc
        return self.scrape_ask_max_pages()

    def scrape_ask_max_pages(self):
        self.state.step = "max_pages"
        rows = [
            [{"text": "1 頁", "callback_data": "mp:1"},
             {"text": "3 頁", "callback_data": "mp:3"},
             {"text": "5 頁", "callback_data": "mp:5"}],
            [{"text": "10 頁", "callback_data": "mp:10"},
             {"text": "全部", "callback_data": "mp:0"}],
        ]
        if self.last.get("max_pages") is not None:
            rows.insert(0, [{
                "text": f"↺ 同上次 ({self.last['max_pages']} 頁)",
                "callback_data": f"mp:{self.last['max_pages']}",
            }])
        self.reply_buttons("幾多頁? (或者自己打數字)", rows)

    def scrape_ask_threshold(self):
        self.state.step = "threshold"
        rows = [
            [{"text": "0 (全部)", "callback_data": "th:0"},
             {"text": "30", "callback_data": "th:30"},
             {"text": "50", "callback_data": "th:50"},
             {"text": "70", "callback_data": "th:70"}],
        ]
        last_th = self.last.get("match_threshold")
        if last_th is not None:
            rows.insert(0, [{
                "text": f"↺ 同上次 ({last_th})",
                "callback_data": f"th:{last_th}",
            }])
        self.reply_buttons(
            "CV match 門檻 (0=全部 push, 高=只 push 啱嘅;\n"
            "或者自己打 0-100 數字):",
            rows,
        )

    def scrape_ask_max_telegram(self):
        self.state.step = "max_telegram"
        rows = [
            [{"text": "5", "callback_data": "mt:5"},
             {"text": "10", "callback_data": "mt:10"},
             {"text": "20", "callback_data": "mt:20"}],
            [{"text": "50", "callback_data": "mt:50"},
             {"text": "0 (無限)", "callback_data": "mt:0"}],
        ]
        last_mt = self.last.get("telegram_max")
        if last_mt is not None:
            rows.insert(0, [{
                "text": f"↺ 同上次 ({last_mt or '無限'})",
                "callback_data": f"mt:{last_mt}",
            }])
        self.reply_buttons(
            "最多 push 幾多張 Telegram 卡? (0 = 無限;\n"
            "Master/CSV 仍記錄全部，呢個只限 Telegram):",
            rows,
        )

    def scrape_show_confirm(self):
        d = self.state.data
        if not d.get("source") or not d.get("keyword"):
            self.state.reset()
            return self.reply(
                "⚠ 缺少 source 或 keyword（flow 中斷過）。\n"
                "打 /scrape 重新開始。"
            )
        self.state.step = "confirm"
        loc_disp = d.get("location") or "(全港)"
        mp = d.get("max_pages")
        if mp is None:
            mp = self.last.get("max_pages", 5)
        if mp == 0:
            mp_disp = "全部"
        else:
            mp_disp = f"{mp} 頁"
        th = d.get("match_threshold")
        if th is None:
            th = self.last.get("match_threshold", 0)
        mt = d.get("telegram_max")
        if mt is None:
            mt = self.last.get("telegram_max", 0)
        mt_disp = "無限" if mt == 0 else f"{mt} 張"
        msg = (
            "<b>確認設定:</b>\n"
            f"🌐 Source: <b>{d['source']}</b>\n"
            f"🔍 Keyword: <b>{d['keyword']}</b>\n"
            f"📍 Location: <b>{loc_disp}</b>\n"
            f"📄 Pages: <b>{mp_disp}</b>\n"
            f"🎯 Threshold: <b>{th}</b>\n"
            f"📱 Max Telegram: <b>{mt_disp}</b>\n\n"
            "新 job 會 push 卡片過嚟，全部 jobs 寫入 Master。\n"
            "每張卡有 [⭐ Save] [🚫 Hide] [✅ Applied] 按掣。"
        )
        self.reply_buttons(msg, INLINE_CONFIRM)

    def scrape_start_now(self):
        d = self.state.data
        self.state.reset()
        if self.scrape_thread and self.scrape_thread.is_alive():
            return self.reply("⚠️ 已經有 scrape 跑緊。")

        # Resolve max_pages / threshold / max_telegram (fall back to last-used)
        mp = d.get("max_pages")
        if mp is None:
            mp = self.last.get("max_pages", 5)
        th = d.get("match_threshold")
        if th is None:
            th = self.last.get("match_threshold", 0)
        mt = d.get("telegram_max")
        if mt is None:
            mt = self.last.get("telegram_max", 0)

        # Persist this run's choices for next time
        self.last.update({
            "source": d["source"],
            "location": d.get("location", ""),
            "max_pages": mp,
            "match_threshold": th,
            "telegram_max": mt,
        })
        save_state(self.persist)

        ns = argparse.Namespace(
            source=d["source"],
            keyword=d["keyword"],
            location=d.get("location", ""),
            max_pages=int(mp),
            full_jd=True,
            delay=1.5,
            output=None,
            csv=False,
            master=self.master_path,
            telegram_enabled=True,
            telegram_token=self.token,
            telegram_chat_id=self.chat_id,
            telegram_max=int(mt),
            telegram_delay=1.5,
            cv=self.cv_path,
            match_threshold=float(th),
            include_actions=True,
            at=None,
        )
        self.scrape_stop = threading.Event()
        self.scrape_args = ns
        bot_self = self

        def worker():
            try:
                scraper.scrape(ns, stop_event=bot_self.scrape_stop)
                bot_self.reply(
                    f"✅ 完成: {ns.source} / {ns.keyword}"
                    + (f" / {ns.location}" if ns.location else "")
                )
            except Exception as e:
                traceback.print_exc()
                bot_self.reply(f"❌ Scrape 失敗: {e}")
            finally:
                bot_self.scrape_thread = None
                bot_self.scrape_args = None

        self.scrape_thread = threading.Thread(target=worker, daemon=True)
        self.scrape_thread.start()
        self.reply(
            f"🚀 開始: <b>{d['source']}</b> / <b>{d['keyword']}</b>"
            + (f" / {d['location']}" if d.get("location") else "")
            + "\n每爬到一個新 job 都會推張卡片落呢度。"
        )

    def cmd_stop(self):
        if self.scrape_thread and self.scrape_thread.is_alive():
            self.scrape_stop.set()
            self.reply("🛑 已發 stop signal，等而家果個 request 完先停。")
        else:
            self.reply("ℹ 而家冇 scrape 跑緊。")

    def cmd_status(self):
        if self.scrape_thread and self.scrape_thread.is_alive() and self.scrape_args:
            ns = self.scrape_args
            self.reply(
                f"🟢 跑緊: <b>{ns.source}</b> / <b>{ns.keyword}</b>"
                + (f" / {ns.location}" if ns.location else "")
            )
        else:
            self.reply("🟡 Idle. 打 /scrape 開始爬。")

    # ---------- /find ----------

    def cmd_find(self, text):
        parts = text.split(maxsplit=1)
        if len(parts) >= 2:
            return self.find_run(parts[1].strip())
        self.state.flow = "find"
        self.state.step = "keyword"
        self.reply("🔍 打你想搵嘅 keyword (job title / company / requirements 都會搜):")

    def find_handle_text(self, text):
        kw = text.strip()
        self.state.reset()
        self.find_run(kw)

    def find_run(self, kw):
        headers, rows = self._master_rows()
        if rows is None:
            return
        try:
            c_score = headers.index("Match Score")
        except ValueError:
            c_score = -1
        kw_low = kw.lower()
        matches = []
        for r in rows:
            blob = " ".join(str(v) for v in r if v is not None).lower()
            if kw_low in blob:
                matches.append(r)
        if c_score >= 0:
            def keyf(r):
                v = r[c_score] if c_score < len(r) else 0
                try:
                    return -int(v or 0)
                except (TypeError, ValueError):
                    return 0
            matches.sort(key=keyf)

        if not matches:
            return self.reply(f"🔍 冇結果 for <i>{kw}</i>")

        out = [f"🔍 <b>{len(matches)} 個結果</b> for <i>{kw}</i>"]
        for r in matches[:5]:
            out.append(self._fmt_row_line(headers, r))
        if len(matches) > 5:
            out.append(f"\n... +{len(matches) - 5} more")
        self.reply("\n".join(out))

    # ---------- /last ----------

    def cmd_last(self, text):
        n = self._parse_n(text, default=5, max_n=20)
        headers, rows = self._master_rows()
        if rows is None:
            return
        try:
            c_sa = headers.index("Scraped At")
        except ValueError:
            c_sa = 1
        rows.sort(key=lambda r: str(r[c_sa] or ""), reverse=True)
        out = [f"📋 <b>最新 {min(n, len(rows))} 個</b>"]
        for r in rows[:n]:
            out.append(self._fmt_row_line(headers, r))
        self.reply("\n".join(out))

    # ---------- /top ----------

    def cmd_top(self, text):
        n = self._parse_n(text, default=5, max_n=20)
        headers, rows = self._master_rows()
        if rows is None:
            return
        if "Match Score" not in headers:
            return self.reply("Master 冇 Match Score 欄。")
        c_score = headers.index("Match Score")

        def keyf(r):
            v = r[c_score] if c_score < len(r) else 0
            try:
                return -int(v or 0)
            except (TypeError, ValueError):
                return 0
        rows.sort(key=keyf)

        out = [f"🏆 <b>Top {min(n, len(rows))} by match</b>"]
        for r in rows[:n]:
            out.append(self._fmt_row_line(headers, r, show_score=True))
        self.reply("\n".join(out))

    # ---------- /stats ----------

    def cmd_stats(self):
        headers, rows = self._master_rows()
        if rows is None:
            return
        c_src = headers.index("Source") if "Source" in headers else 0
        c_sa = headers.index("Scraped At") if "Scraped At" in headers else 1
        c_score = headers.index("Match Score") if "Match Score" in headers else -1

        src_counts = Counter(r[c_src] for r in rows if r[c_src])
        scraped_dates = [str(r[c_sa])[:10] for r in rows if r[c_sa]]
        latest = max(scraped_dates) if scraped_dates else "?"
        earliest = min(scraped_dates) if scraped_dates else "?"

        high_match = 0
        if c_score >= 0:
            for r in rows:
                try:
                    if int(r[c_score] or 0) >= 70:
                        high_match += 1
                except (TypeError, ValueError):
                    pass

        out = [
            "📊 <b>Master stats</b>",
            f"總數: <b>{len(rows)}</b> jobs",
            "",
            "By source:",
        ]
        for src, n in src_counts.most_common():
            out.append(f"  • {src}: {n}")
        out.append("")
        out.append(f"📅 日期: {earliest} → {latest}")
        if c_score >= 0:
            out.append(f"🎯 70%+ match: {high_match}")
        self.reply("\n".join(out))

    # ---------- formatting helpers ----------

    def _parse_n(self, text, default, max_n):
        parts = text.split()
        if len(parts) >= 2:
            try:
                return max(1, min(max_n, int(parts[1])))
            except ValueError:
                pass
        return default

    def _fmt_row_line(self, headers, row, show_score=False):
        d = dict(zip(headers, row))
        title = (str(d.get("Job Title") or "(no title)"))[:60]
        company = (str(d.get("Company") or ""))[:40]
        url = (d.get("URL") or "")
        src = d.get("Source") or ""
        score = d.get("Match Score") or ""
        line = "\n• "
        if show_score and score:
            line += f"🎯 <b>{score}%</b> "
        line += f"<b>{self._esc(title)}</b>"
        if company:
            line += f" — {self._esc(company)}"
        line += f"\n   {self._esc(src)}"
        if score and not show_score:
            line += f" · {score}% match"
        if url:
            line += f" · <a href=\"{self._esc(url)}\">link</a>"
        return line

    @staticmethod
    def _esc(s):
        return (
            str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        )


def load_config():
    if not CONFIG_PATH.exists():
        return {}
    try:
        return json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {}


def main():
    p = argparse.ArgumentParser(description="Telegram bot listener")
    p.add_argument("--token", default="")
    p.add_argument("--chat-id", default="")
    p.add_argument("--master", default="")
    p.add_argument("--cv", default="")
    args = p.parse_args()

    cfg = load_config()
    token = args.token or cfg.get("tg_token", "")
    chat_id = args.chat_id or cfg.get("tg_chat", "")
    master = args.master or cfg.get("master", "") or str(APP_DIR / "jobs_master.xlsx")
    cv = args.cv or cfg.get("cv", "")

    if not token or not chat_id:
        print("ERROR: bot token + chat ID required.")
        print(
            "Set them in GUI (saved to config.json), or pass --token / --chat-id."
        )
        return 2

    bot = Bot(token=token, chat_id=chat_id, master_path=master, cv_path=cv)
    try:
        bot.run()
    except KeyboardInterrupt:
        print("\nBot stopped by Ctrl+C.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
