"""Sync /tmp/jobs_master.xlsx ↔ Supabase master_jobs table.

Why both: scraper.py expects a local xlsx file (MasterDB class operates
on it for dedup + new-row appends). On Streamlit Cloud /tmp is
ephemeral — wiped on each container restart, so we lose accumulated
jobs. The fix: at scrape start, rebuild the xlsx from Supabase; at
scrape end, upsert new rows back to Supabase. Scraper.py untouched.

Public API used by streamlit_app.py:
    fetch_master_rows(supabase, user_id) -> list[dict]
    download_master_to_xlsx(supabase, user_id, xlsx_path) -> int
    sync_xlsx_to_supabase(supabase, user_id, xlsx_path) -> (added, updated)
    build_xlsx_bytes(supabase, user_id) -> bytes      # for download button
"""

import io
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None


# Mirrors scraper.FIELDNAMES_MASTER ordering (Source/Scraped At first,
# then the regular fields, then status flags Saved/Hidden/Applied).
FIELDNAMES_MASTER = [
    "Source", "Scraped At",
    "JD Number", "Job Title", "Company", "Salary", "Location",
    "Posted Date", "Posted (display)", "Classification", "Work Type",
    "Responsibilities", "Requirements", "Benefits", "How to apply",
    "URL", "Match Score", "Match Keywords",
    "Saved", "Hidden", "Applied",
]

# xlsx header → Supabase column (snake_case)
XLSX_TO_DB = {
    "Source":            "source",
    "JD Number":         "jd_number",
    "Job Title":         "job_title",
    "Company":           "company",
    "Salary":            "salary",
    "Location":          "location",
    "Posted Date":       "posted_date",
    "Posted (display)":  "posted_display",
    "Classification":    "classification",
    "Work Type":         "work_type",
    "Responsibilities":  "responsibilities",
    "Requirements":      "requirements",
    "Benefits":          "benefits",
    "How to apply":      "how_to_apply",
    "URL":               "url",
    "Match Score":       "match_score",
    "Match Keywords":    "match_keywords",
}

DB_TO_XLSX = {v: k for k, v in XLSX_TO_DB.items()}


# ============================================================
# Read from Supabase
# ============================================================

def fetch_master_rows(supabase, user_id, limit=None):
    """Return all master_jobs rows for the user.

    `limit` is None → fetch all; otherwise paginate with .range() to
    avoid Supabase's default 1000-row cap.
    """
    if not supabase or not user_id:
        return []
    rows = []
    page_size = 1000
    offset = 0
    while True:
        q = (
            supabase.table("master_jobs")
            .select("*")
            .eq("user_id", str(user_id))
            .order("scraped_at", desc=True)
            .range(offset, offset + page_size - 1)
        )
        try:
            res = q.execute()
        except Exception as e:
            print(f"  [master-sync] fetch_master_rows page failed: {e}")
            break
        batch = res.data or []
        rows.extend(batch)
        if len(batch) < page_size:
            break
        offset += page_size
        if limit and len(rows) >= limit:
            rows = rows[:limit]
            break
    return rows


def _ws_append_row(ws, db_row):
    """Append a master_jobs row (dict, snake_case) onto an xlsx worksheet,
    laying out columns in FIELDNAMES_MASTER order."""
    out = []
    for h in FIELDNAMES_MASTER:
        if h == "Scraped At":
            out.append(db_row.get("scraped_at") or "")
        elif h in ("Saved", "Hidden", "Applied"):
            # status flags don't live on master_jobs — left blank here;
            # they're merged in when needed (e.g. building UI views).
            out.append("")
        else:
            col = XLSX_TO_DB.get(h)
            v = db_row.get(col) if col else None
            out.append(v if v is not None else "")
    ws.append(out)


def download_master_to_xlsx(supabase, user_id, xlsx_path):
    """Rebuild a local xlsx from Supabase rows. Returns row count written.

    Overwrites any existing xlsx at `xlsx_path` — scraper's MasterDB
    will read this on its next .open() call.
    """
    if openpyxl is None:
        print("  [master-sync] openpyxl not installed; skipping pre-populate")
        return 0
    rows = fetch_master_rows(supabase, user_id)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Jobs"
    ws.append(FIELDNAMES_MASTER)
    for r in rows:
        _ws_append_row(ws, r)
    Path(xlsx_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
    return len(rows)


def build_xlsx_bytes(supabase, user_id):
    """Build an in-memory xlsx of the user's full master. Returns bytes."""
    if openpyxl is None:
        return b""
    rows = fetch_master_rows(supabase, user_id)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Jobs"
    ws.append(FIELDNAMES_MASTER)
    for r in rows:
        _ws_append_row(ws, r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================================================
# Write to Supabase (post-scrape)
# ============================================================

def _coerce(v):
    """Convert blanks to None and trim strings before inserting."""
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    return v


def sync_xlsx_to_supabase(supabase, user_id, xlsx_path):
    """Read xlsx and upsert rows to Supabase. Returns (added, updated)."""
    if openpyxl is None or not supabase or not user_id:
        return 0, 0
    p = Path(xlsx_path)
    if not p.exists():
        return 0, 0
    try:
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    except Exception as e:
        print(f"  [master-sync] could not open {p}: {e}")
        return 0, 0
    ws = wb.active
    if ws.max_row is None or ws.max_row < 2:
        wb.close()
        return 0, 0
    headers = [c.value for c in ws[1]]

    # Snapshot existing JD numbers so we can report added vs updated
    try:
        existing_res = (
            supabase.table("master_jobs")
            .select("jd_number")
            .eq("user_id", str(user_id))
            .execute()
        )
        existing = {r["jd_number"] for r in (existing_res.data or [])}
    except Exception as e:
        print(f"  [master-sync] could not fetch existing IDs: {e}")
        existing = set()

    to_upsert = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        jd = (record.get("JD Number") or "").strip() if record.get("JD Number") else ""
        if not jd:
            continue
        db_row = {"user_id": str(user_id), "jd_number": jd}
        for h, col in XLSX_TO_DB.items():
            if h == "JD Number":
                continue
            v = _coerce(record.get(h))
            if v is not None:
                # match_score should be numeric
                if col == "match_score":
                    try:
                        v = float(v)
                    except (TypeError, ValueError):
                        continue
                db_row[col] = v
        if record.get("Scraped At"):
            db_row["scraped_at"] = str(record["Scraped At"])
        to_upsert.append(db_row)
    wb.close()

    if not to_upsert:
        return 0, 0

    # Batch upsert (Supabase has body-size limit; 500 is conservative)
    BATCH = 500
    for i in range(0, len(to_upsert), BATCH):
        chunk = to_upsert[i : i + BATCH]
        try:
            supabase.table("master_jobs").upsert(chunk).execute()
        except Exception as e:
            print(f"  [master-sync] upsert batch {i // BATCH} failed: {e}")
            return 0, 0

    added = sum(1 for r in to_upsert if r["jd_number"] not in existing)
    updated = len(to_upsert) - added
    return added, updated
