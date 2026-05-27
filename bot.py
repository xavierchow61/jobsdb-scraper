"""Telegram bot listener: trigger scrapes + query master DB via chat commands.

Reads telegram token and master path from config.json (the GUI writes this).
Run with: python bot.py   (or double-click Bot.bat)
Stops with Ctrl+C.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import threading
import time
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime
from pathlib import Path

import scraper

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except (AttributeError, OSError):
    pass

APP_DIR = Path(__file__).resolve().parent
CONFIG_PATH = APP_DIR / "config.json"
TELEGRAM_API = "https://api.telegram.org"


def load_config():
    if not CONFIG_PATH.exists():
        return {}
    try:
        return json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {}


class Bot:
    def __init__(self, token, default_chat=None, allowed_chats=None):
        self.token = token
        self.default_chat = str(default_chat) if default_chat else ""
        self.allowed = set(str(c) for c in (allowed_chats or []) if c)
        self.offset = None
        self.scrape_stop = threading.Event()
        self.scrape_thread = None
        self.stop = threading.Event()

    # --- HTTP helpers ---

    def _call(self, method, payload):
        url = f"{TELEGRAM_API}/bot{self.token}/{method}"
        data = json.dumps(payload).encode("utf-8")
        req = urllib.request.Request(
            url, data=data, method="POST",
            headers={"Content-Type": "application/json"},
        )
        try:
            with urllib.request.urlopen(req, timeout=60) as resp:
                return json.loads(resp.read().decode("utf-8"))
        except urllib.error.HTTPError as e:
            return {"ok": False, "error_code": e.code,
                    "description": e.read().decode("utf-8", errors="replace")[:400]}
        except Exception as e:
            return {"ok": False, "description": str(e)}

    def send(self, chat_id, text, parse_mode="HTML"):
        return self._call("sendMessage", {
            "chat_id": str(chat_id),
            "text": text,
            "parse_mode": parse_mode,
            "disable_web_page_preview": True,
        })

    # --- Main loop ---

    def listen(self):
        print(f"Bot listening. Allowed chats: "
              f"{self.allowed if self.allowed else '(any chat)'}")
        print("Send /help to your bot in Telegram. Press Ctrl+C to stop.\n")
        while not self.stop.is_set():
            params = {"timeout": 30, "allowed_updates": ["message"]}
            if self.offset is not None:
                params["offset"] = self.offset
            resp = self._call("getUpdates", params)
            if not resp.get("ok"):
                desc = resp.get("description", "?")
                print(f"  poll error: {desc[:150]}")
                if self.stop.wait(5):
                    break
                continue
            for upd in resp.get("result", []):
                self.offset = upd["update_id"] + 1
                try:
                    self.handle(upd)
                except Exception as e:
                    print(f"  handler error: {e}")

    def handle(self, update):
        msg = update.get("message") or update.get("edited_message") or {}
        chat = msg.get("chat") or {}
        chat_id = str(chat.get("id", ""))
        text = (msg.get("text") or "").strip()
        if not text or not chat_id:
            return

        if self.allowed and chat_id not in self.allowed:
            print(f"  ignoring message from chat {chat_id}: {text[:50]}")
            self.send(chat_id, "Unauthorized")
            return

        print(f"  [{datetime.now():%H:%M:%S}] chat {chat_id}: {text[:80]}")

        cmd, _, rest = text.partition(" ")
        cmd = cmd.lower().lstrip("/")
        rest = rest.strip()

        handlers = {
            "help": self.cmd_help,
            "start": self.cmd_help,
            "scrape": self.cmd_scrape,
            "query": self.cmd_query,
            "stats": self.cmd_stats,
            "cancel": self.cmd_cancel,
            "stop": self.cmd_cancel,
            "threshold": self.cmd_threshold,
        }
        h = handlers.get(cmd)
        if h is None:
            self.send(chat_id, f"Unknown command: /{cmd}\nSend /help for a list.")
            return
        h(chat_id, rest)

    # --- Commands ---

    def cmd_help(self, chat_id, _rest):
        msg = (
            "<b>🤖 HK Job Scraper Bot</b>\n\n"
            "<b>/scrape</b> SOURCE KEYWORD [LOCATION]\n"
            "  Quick scrape (≤ 2 pages), live-push each new job.\n"
            "  SOURCE = <code>jobsdb</code> / <code>ctgoodjobs</code> / <code>cpjobs</code>\n"
            "  e.g. <code>/scrape cpjobs accountant Kowloon</code>\n"
            "  e.g. <code>/scrape ctgoodjobs accounting Tseung Kwan O</code>\n\n"
            "<b>/query</b> TEXT\n"
            "  Search Master Excel; returns top 5 matches.\n"
            "  e.g. <code>/query AR Excel 5 years</code>\n\n"
            "<b>/stats</b> — totals per source\n"
            "<b>/cancel</b> — stop the running scrape\n"
            "<b>/threshold</b> N — set match score threshold for live pushes\n"
            "<b>/help</b> — this message"
        )
        self.send(chat_id, msg)

    def cmd_scrape(self, chat_id, rest):
        if self.scrape_thread and self.scrape_thread.is_alive():
            self.send(chat_id, "⚠ A scrape is already running. Use /cancel first.")
            return
        if not rest:
            self.send(chat_id, "Usage: /scrape SOURCE KEYWORD [LOCATION]")
            return

        parts = rest.split(maxsplit=2)
        source = parts[0].lower() if parts else ""
        if source not in scraper.SOURCES:
            self.send(
                chat_id,
                f"Unknown source: <code>{source}</code>\n"
                f"Valid: {', '.join(scraper.SOURCES)}",
            )
            return
        if len(parts) < 2:
            self.send(chat_id, "Usage: /scrape SOURCE KEYWORD [LOCATION]")
            return
        keyword = parts[1]
        location = parts[2] if len(parts) >= 3 else ""

        config = load_config()
        master = config.get("master") or str(APP_DIR / "jobs_master.xlsx")
        cv = config.get("cv") or ""
        try:
            threshold = float(config.get("match_threshold") or 0)
        except (TypeError, ValueError):
            threshold = 0.0

        ns = argparse.Namespace(
            source=source,
            keyword=keyword,
            location=location,
            max_pages=2,
            full_jd=True,
            delay=1.5,
            output=None,
            csv=False,
            master=master,
            telegram_enabled=True,
            telegram_token=self.token,
            telegram_chat_id=chat_id,
            telegram_max=0,
            telegram_delay=1.5,
            cv=cv,
            match_threshold=threshold,
            at=None,
        )

        self.send(
            chat_id,
            f"🔍 Starting <b>{source}</b> scrape\n"
            f"  keyword: <code>{keyword}</code>\n"
            f"  location: <code>{location or '(any)'}</code>\n"
            f"  threshold: {threshold:.0f}%\n"
            f"  pages: 2 (use GUI for full scan)",
        )

        self.scrape_stop = threading.Event()
        stop_event = self.scrape_stop

        def runner():
            try:
                scraper.scrape(ns, stop_event=stop_event)
                self.send(chat_id, "✅ Scrape finished.")
            except Exception as e:
                self.send(chat_id, f"❌ Scrape error: {e}")

        self.scrape_thread = threading.Thread(
            target=runner, name="bot-scrape", daemon=True
        )
        self.scrape_thread.start()

    def cmd_cancel(self, chat_id, _rest):
        if self.scrape_thread and self.scrape_thread.is_alive():
            self.scrape_stop.set()
            self.send(chat_id, "🛑 Stop requested; finishing current request…")
        else:
            self.send(chat_id, "No scrape is running.")

    def cmd_threshold(self, chat_id, rest):
        try:
            n = float(rest)
        except (ValueError, TypeError):
            self.send(chat_id, "Usage: /threshold 50  (0–100)")
            return
        n = max(0.0, min(100.0, n))
        config = load_config()
        config["match_threshold"] = n
        try:
            CONFIG_PATH.write_text(
                json.dumps(config, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            self.send(chat_id, f"✅ Threshold set to {n:.0f}% (saved to config).")
        except Exception as e:
            self.send(chat_id, f"❌ Could not save config: {e}")

    def cmd_query(self, chat_id, rest):
        if not rest:
            self.send(chat_id, "Usage: /query TEXT")
            return
        if load_workbook is None:
            self.send(chat_id, "openpyxl not installed — cannot read master.")
            return

        config = load_config()
        master = Path(config.get("master") or APP_DIR / "jobs_master.xlsx")
        if not master.exists():
            self.send(chat_id, f"Master file not found: <code>{master}</code>")
            return

        try:
            wb = load_workbook(master, read_only=True, data_only=True)
        except Exception as e:
            self.send(chat_id, f"Cannot open master: {e}")
            return
        ws = wb.active
        headers = [c.value for c in ws[1]]
        col = {h: i for i, h in enumerate(headers) if h}

        keywords = [k.lower() for k in re.findall(r"\S+", rest)]
        search_cols = [
            col[k] for k in
            ("Job Title", "Company", "Classification",
             "Responsibilities", "Requirements", "Benefits")
            if k in col
        ]

        matches = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            haystack = " ".join(
                str(row[i] or "").lower() for i in search_cols if i < len(row)
            )
            if all(k in haystack for k in keywords):
                matches.append(row)
        wb.close()

        if not matches:
            self.send(chat_id, f"No matches for: <code>{rest}</code>")
            return

        if "Match Score" in col:
            ms = col["Match Score"]

            def _score(r):
                try:
                    return int(r[ms]) if r[ms] is not None else 0
                except (TypeError, ValueError):
                    return 0
            matches.sort(key=_score, reverse=True)

        n_show = min(5, len(matches))
        self.send(
            chat_id,
            f"📊 Found {len(matches)} match(es) for "
            f"<code>{rest}</code>; showing top {n_show}",
        )

        for r in matches[:n_show]:
            row_dict = {h: r[col[h]] for h in headers if col.get(h, -1) < len(r) and h}
            src = row_dict.get("Source") or "?"
            ms_val = row_dict.get("Match Score")
            if isinstance(ms_val, str) and ms_val.isdigit():
                row_dict["Match Score"] = int(ms_val)
            text = scraper.format_telegram_card(row_dict, src)
            self.send(chat_id, text)
            time.sleep(1.5)

    def cmd_stats(self, chat_id, _rest):
        if load_workbook is None:
            self.send(chat_id, "openpyxl not installed.")
            return
        config = load_config()
        master = Path(config.get("master") or APP_DIR / "jobs_master.xlsx")
        if not master.exists():
            self.send(chat_id, f"Master file not found: <code>{master}</code>")
            return
        try:
            wb = load_workbook(master, read_only=True, data_only=True)
        except Exception as e:
            self.send(chat_id, f"Cannot open master: {e}")
            return
        ws = wb.active
        headers = [c.value for c in ws[1]]
        col = {h: i for i, h in enumerate(headers) if h}
        src_idx = col.get("Source", 0)
        score_idx = col.get("Match Score")

        counts = {}
        score_sum = score_count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            s = str(row[src_idx] or "?") if src_idx < len(row) else "?"
            counts[s] = counts.get(s, 0) + 1
            if score_idx is not None and score_idx < len(row):
                v = row[score_idx]
                try:
                    iv = int(v) if v is not None else None
                except (TypeError, ValueError):
                    iv = None
                if iv is not None and iv > 0:
                    score_sum += iv
                    score_count += 1
        wb.close()
        total = sum(counts.values())
        lines = [f"<b>📊 Master DB</b>", f"Total: {total}"]
        for src, n in sorted(counts.items(), key=lambda x: -x[1]):
            lines.append(f"  • <b>{src}</b>: {n}")
        if score_count:
            lines.append(f"\nAvg Match Score: {score_sum / score_count:.1f}% "
                         f"(scored: {score_count})")
        self.send(chat_id, "\n".join(lines))


def main():
    p = argparse.ArgumentParser(description="Telegram bot for HK Job Scraper")
    p.add_argument("--token", default="",
                   help="Telegram bot token (else from config.json)")
    p.add_argument("--chat", default="",
                   help="Chat ID allowed to send commands (else from config.json)")
    p.add_argument("--any-chat", action="store_true",
                   help="Accept commands from any chat (default: only configured chat)")
    args = p.parse_args()

    config = load_config()
    token = args.token or config.get("tg_token") or ""
    chat = args.chat or config.get("tg_chat") or ""
    if not token:
        print("ERROR: No Telegram bot token. Set it in the GUI Settings, "
              "or pass --token.")
        sys.exit(2)

    allowed = [] if args.any_chat else ([chat] if chat else [])
    bot = Bot(token, default_chat=chat, allowed_chats=allowed)
    try:
        bot.listen()
    except KeyboardInterrupt:
        print("\nStopped by Ctrl+C.")
        bot.stop.set()


if __name__ == "__main__":
    main()
