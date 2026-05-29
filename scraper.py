"""HK job scraper - JobsDB / CTgoodjobs / cpjobs -> master Excel + Telegram."""

import argparse
import csv
import html as html_mod
import json
import re
import sys
import time
import urllib.error
import urllib.request
from datetime import datetime, timedelta
from pathlib import Path

from curl_cffi import requests

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    Workbook = None
    load_workbook = None

try:
    import cv_match
except ImportError:
    cv_match = None

try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except (AttributeError, OSError):
    pass

BASE = "https://hk.jobsdb.com"
CT_BASE = "https://jobs.ctgoodjobs.hk"
CP_API = "https://api.cpjobs.com"
CP_BASE = "https://www.cpjobs.com/hk"
IMPERSONATE = "chrome124"

SOURCES = ("jobsdb", "ctgoodjobs", "cpjobs")

# cpjobs only exposes 4 top-level HK regions (no sub-districts)
CP_LOCATION_MAP = {
    "Hong Kong Island": "C2",
    "Kowloon": "C3",
    "New Territories": "C4",
    "Outlying Islands": "C5",
}
CP_LOCATIONS = list(CP_LOCATION_MAP)
CT_HK_REGIONS = ("hong-kong", "kowloon", "new-territories", "outlying-islands")

# Known HK districts on CTgoodjobs (harvested from their site).
# Used by the GUI to populate a Location dropdown for the CT source.
CT_LOCATIONS = [
    "Aberdeen", "Admiralty", "Ap Lei Chau", "Causeway Bay", "Central",
    "Central and Western District", "Chai Wan", "Chek Lap Kok", "Cheung Chau",
    "Cheung Sha Wan", "Choi Hung", "Chung Hom Kok", "Clear Water Bay",
    "Cyberport", "Diamond Hill", "Discovery Bay", "Eastern District",
    "Fanling", "Fo Tan", "Fortress Hill", "HK International Airport",
    "Happy Valley", "Ho Man Tin", "Hong Kong Science Park", "Hung Hom",
    "Islands District", "Jordan", "Kai Tak", "Kam Tin", "Kennedy Town",
    "Kowloon Bay", "Kowloon City", "Kowloon City District", "Kowloon Tong",
    "Kwai Chung", "Kwai Fong", "Kwai Hing", "Kwai Tsing District", "Kwun Tong",
    "Kwun Tong District", "Lai Chi Kok", "Lai King", "Lam Tin", "Lamma Island",
    "Lantau", "Lo Wu", "Lok Fu", "Lok Ma Chau", "Ma On Shan", "Ma Wan",
    "Mei Foo", "Mid Level", "Mong Kok", "Morrison Hill", "Ngau Chi Wan",
    "Ngau Tau Kok", "North Point", "Northern District", "Pak Shek Kok",
    "Pat Heung", "Pok Fu Lam", "Prince Edward", "Quarry Bay", "Sai Kung",
    "Sai Kung District", "Sai Wan Ho", "Sai Ying Pun", "San Po Kong",
    "Sha Tin", "Sha Tin District", "Sham Shui Po", "Sham Shui Po District",
    "Shau Kei Wan", "Shek Kip Mei", "Shek Kong", "Shek Mun", "Sheung Shui",
    "Sheung Wan", "Siu Lek Yuen", "Siu Sai Wan", "Southern District",
    "Stanley", "Tai Kok Tsui", "Tai Po", "Tai Po District", "Tai Tam",
    "Tai Wai", "Taikoo Shing", "The Peak", "Tin Hau", "Tin Shui Wai",
    "To Kwa Wan", "Tseung Kwan O", "Tsim Sha Tsui", "Tsim Sha Tsui East",
    "Tsing Yi", "Tsuen Wan", "Tsuen Wan District", "Tsz Wan Shan",
    "Tuen Mun", "Tuen Mun District", "Tung Chung", "Wan Chai",
    "Wan Chai District", "Wong Chuk Hang", "Wong Tai Sin",
    "Wong Tai Sin District", "Yau Ma Tei", "Yau Tong",
    "Yau Tsim Mong District", "Yuen Long", "Yuen Long District",
]
HEADERS = {
    "Accept": (
        "text/html,application/xhtml+xml,application/xml;q=0.9,"
        "image/avif,image/webp,image/apng,*/*;q=0.8"
    ),
    "Accept-Language": "en-HK,en;q=0.9,zh-HK;q=0.8",
    "Upgrade-Insecure-Requests": "1",
}

FIELDNAMES = [
    "JD Number",
    "Job Title",
    "Company",
    "Salary",
    "Location",
    "Posted Date",
    "Posted (display)",
    "Classification",
    "Work Type",
    "Responsibilities",
    "Requirements",
    "Benefits",
    "How to apply",
    "URL",
    "Match Score",
    "Match Keywords",
    "Saved",
    "Hidden",
    "Applied",
]
# Extra columns added by the master Excel file
FIELDNAMES_MASTER = ["Source", "Scraped At"] + FIELDNAMES

DEFAULT_MASTER_FILENAME = "jobs_master.xlsx"
TELEGRAM_API = "https://api.telegram.org"

RESP_PATTERN = re.compile(
    r"(responsibilit|duties|the\s+role|your\s+role|"
    r"what\s+you('?ll|\s+will)\s+do|key\s+tasks|job\s+function|"
    r"main\s+(duties|responsibilities)|primary\s+responsibilities|"
    r"職\s*責|工作\s*內容|主要\s*工作|工作\s*範圍|職務)",
    re.IGNORECASE,
)
REQ_PATTERN = re.compile(
    r"(requirement|qualification|what\s+we('?re|\s+are)\s+looking|"
    r"about\s+you|candidate\s+profile|ideal\s+candidate|who\s+you\s+are|"
    r"you('?ll|\s+will)\s+need|you\s+(should|must)\s+have|"
    r"skills?\s+(needed|required|and\s+experience)|experience\s+required|"
    r"要\s*求|資\s*格|入職\s*要求|條件|資歷|應徵\s*條件)",
    re.IGNORECASE,
)
APPLY_PATTERN = re.compile(
    r"(how\s+to\s+apply|to\s+apply|application\s+(method|procedure|details?)|"
    r"apply\s+(now|online|here|via|by|with|through)|"
    r"interested\s+(parties|candidates|applicants)|"
    r"please\s+(send|submit|email|forward|apply)|"
    r"contact\s+(us|details|information|person)|"
    r"申請\s*方法|聯\s*絡(方法|資料|人)?|連\s*絡|請\s*將|請\s*寄|請\s*致電|"
    r"履歷\s*寄|履歷\s*交|有興趣)",
    re.IGNORECASE,
)
BENEFITS_PATTERN = re.compile(
    r"(^benefits?$|^benefits?[:：]|what\s+we\s+offer|"
    r"we\s+(offer|provide)|what'?s?\s+in\s+it\s+for\s+you|"
    r"why\s+(join|work\s+with|choose)\s+us|"
    r"compensation\s+(and|&)\s+benefit|remuneration\s+package|"
    r"package\s+(includes?|offered)|perks(\s+and\s+benefits)?|"
    r"福\s*利|待\s*遇|薪酬\s*(福利|待遇)?|"
    r"我\s*(們|哋)\s*提供|員工\s*福利|公司\s*福利)",
    re.IGNORECASE,
)
EMAIL_RE = re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")


# ============================================================
# Master Excel (cumulative DB) + Telegram live notifications
# ============================================================

class MasterDB:
    """Append-only master Excel file. Tracks JD Numbers to dedupe across runs."""

    def __init__(self, path):
        self.path = Path(path).resolve() if path else None
        self.wb = None
        self.ws = None
        self.existing_ids = set()
        self.dirty = False

    def open(self):
        if not self.path:
            return
        if load_workbook is None or Workbook is None:
            raise RuntimeError("openpyxl not installed; cannot use master Excel.")
        if self.path.exists():
            try:
                self.wb = load_workbook(self.path)
            except Exception as e:
                raise RuntimeError(f"Cannot read master {self.path}: {e}")
            self.ws = self.wb.active
            self._migrate_headers_if_needed()
            headers = [c.value for c in self.ws[1]]
            try:
                idx = headers.index("JD Number")
            except ValueError:
                idx = None
            if idx is not None:
                for row in self.ws.iter_rows(min_row=2, values_only=True):
                    if idx < len(row) and row[idx]:
                        self.existing_ids.add(str(row[idx]).strip())
            print(
                f"  Master {self.path.name}: "
                f"{len(self.existing_ids)} existing jobs"
            )
        else:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = "Jobs"
            self.ws.append(FIELDNAMES_MASTER)
            self.dirty = True
            print(f"  Master {self.path.name}: creating new file")

    def _migrate_headers_if_needed(self):
        """Append any FIELDNAMES_MASTER columns missing from this xlsx."""
        if self.ws is None:
            return
        current = [c.value for c in self.ws[1]]
        new_count = 0
        for h in FIELDNAMES_MASTER:
            if h not in current:
                self.ws.cell(row=1, column=len(current) + 1, value=h)
                current.append(h)
                new_count += 1
        if new_count:
            self.dirty = True
            print(f"  Master: added {new_count} new column(s) to header")

    def has(self, jid):
        return bool(jid) and str(jid).strip() in self.existing_ids

    def append(self, row, source):
        if not self.path or self.ws is None:
            return
        jid = str(row.get("JD Number", "")).strip()
        if jid:
            self.existing_ids.add(jid)
        out = {f: row.get(f, "") for f in FIELDNAMES}
        out["Source"] = source
        out["Scraped At"] = datetime.now().isoformat(timespec="seconds")
        self.ws.append([out.get(f, "") for f in FIELDNAMES_MASTER])
        self.dirty = True

    def save(self):
        if not (self.wb and self.dirty and self.path):
            return
        try:
            self.wb.save(self.path)
            self.dirty = False
        except PermissionError:
            print(
                f"  WARN: cannot write to {self.path.name} "
                f"(file open in Excel?). Will retry on next save."
            )
        except Exception as e:
            print(f"  WARN: master save failed: {e}")


# ============================================================
# Master query helpers (used by bot_listener)
# ============================================================

def master_rows(master_path):
    """Yield dicts of every row in the master xlsx."""
    p = Path(master_path)
    if not p.exists() or load_workbook is None:
        return
    wb = load_workbook(p, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    for r in ws.iter_rows(min_row=2, values_only=True):
        yield dict(zip(headers, r))
    wb.close()


def master_set_status(master_path, jd_number, column, value=None):
    """Set a status column (Saved / Hidden / Applied) for a single JD Number.

    value defaults to current ISO timestamp; pass "" to clear.
    Returns True if updated, False if JD not found or write failed.
    """
    p = Path(master_path)
    if not p.exists() or load_workbook is None:
        return False
    if value is None:
        value = datetime.now().isoformat(timespec="seconds")
    try:
        wb = load_workbook(p)
    except Exception as e:
        print(f"  master_set_status: cannot open: {e}")
        return False
    ws = wb.active
    headers = [c.value for c in ws[1]]
    if column not in headers or "JD Number" not in headers:
        # Try to add the column at the end
        if column in FIELDNAMES_MASTER and column not in headers:
            ws.cell(row=1, column=len(headers) + 1, value=column)
            headers.append(column)
        else:
            wb.close()
            return False
    jd_col = headers.index("JD Number") + 1
    val_col = headers.index(column) + 1
    target = str(jd_number).strip()
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=jd_col).value
        if v is not None and str(v).strip() == target:
            ws.cell(row=r, column=val_col, value=value)
            try:
                wb.save(p)
                wb.close()
                return True
            except PermissionError:
                wb.close()
                print("  master_set_status: file locked (open in Excel?)")
                return False
    wb.close()
    return False


def master_stats(master_path):
    """Return a summary dict for the master."""
    sources = {}
    saved = applied = hidden = total = 0
    latest = ""
    for r in master_rows(master_path):
        total += 1
        s = r.get("Source") or "?"
        sources[s] = sources.get(s, 0) + 1
        if r.get("Saved"):
            saved += 1
        if r.get("Applied"):
            applied += 1
        if r.get("Hidden"):
            hidden += 1
        sa = str(r.get("Scraped At") or "")
        if sa > latest:
            latest = sa
    return {
        "total": total,
        "sources": sources,
        "saved": saved,
        "applied": applied,
        "hidden": hidden,
        "latest_scrape": latest,
    }


def _summarize_section(text, max_bullets=3, max_chars_each=90):
    """Pick the first few bullet-like lines, trim each."""
    if not text:
        return []
    bullets = []
    seen = set()
    for raw in text.splitlines():
        ln = raw.strip().lstrip("-•*• \t")
        if not ln or len(ln) < 4:
            continue
        if ln.lower() in seen:
            continue
        seen.add(ln.lower())
        if len(ln) > max_chars_each:
            ln = ln[: max_chars_each - 1].rstrip() + "…"
        bullets.append(ln)
        if len(bullets) >= max_bullets:
            break
    if not bullets:
        flat = " ".join(text.split())
        if flat:
            bullets.append(flat[: max_chars_each * 2] + ("…" if len(flat) > max_chars_each * 2 else ""))
    return bullets


def format_telegram_card(row, source):
    e = html_mod.escape
    title = (row.get("Job Title") or "(no title)").strip()
    company = (row.get("Company") or "").strip()
    salary = (row.get("Salary") or "").strip()
    location = (row.get("Location") or "").strip()
    posted = (row.get("Posted Date") or row.get("Posted (display)") or "").strip()
    if "T" in posted[:11]:
        posted = posted[:10]
    work_type = (row.get("Work Type") or "").strip()
    url = (row.get("URL") or "").strip()

    resp_bullets = _summarize_section(row.get("Responsibilities") or "")
    req_bullets = _summarize_section(row.get("Requirements") or "")

    lines = [f"📋 <b>{e(title)}</b>"]

    # Match score (if CV-based scoring was applied)
    score = row.get("Match Score")
    if isinstance(score, (int, float)) and score > 0:
        s = int(score)
        emoji = "🎯" if s >= 70 else "✨" if s >= 50 else "📊"
        match_kw = (row.get("Match Keywords") or "").strip()
        line = f"{emoji} <b>{s}% Match</b>"
        if match_kw:
            top = ", ".join(match_kw.split(", ")[:5])
            line += f"  <i>{e(top)}</i>"
        lines.append(line)

    if company:
        lines.append(f"🏢 {e(company)}")
    if salary:
        lines.append(f"💰 {e(salary)}")
    if location:
        lines.append(f"📍 {e(location)}")
    if work_type:
        lines.append(f"⏱ {e(work_type)}")
    if posted:
        lines.append(f"📅 {e(posted)}")

    if resp_bullets:
        lines.append("\n📝 <b>Responsibilities</b> (摘要)")
        for b in resp_bullets:
            lines.append(f"• {e(b)}")
    if req_bullets:
        lines.append("\n✅ <b>Requirements</b> (摘要)")
        for b in req_bullets:
            lines.append(f"• {e(b)}")

    if resp_bullets or req_bullets:
        lines.append("\n<i>ℹ 完整 Responsibilities / Requirements 請按下方瀏覽網站</i>")

    # If neither section was found, fall back to a single teaser line
    if not resp_bullets and not req_bullets:
        teaser = (row.get("How to apply") or "").strip()
        teaser = re.sub(r"\s+", " ", teaser)
        if teaser:
            snippet = teaser[:220] + ("…" if len(teaser) > 220 else "")
            lines.append(f"\n📝 <i>{e(snippet)}</i>")

    if url:
        lines.append(f"\n🌐 <b>{e(source)}</b>  |  👉 <a href=\"{e(url)}\">瀏覽完整 JD / Apply</a>")
    else:
        lines.append(f"\n🌐 <b>{e(source)}</b>")
    return "\n".join(lines)


def build_job_action_buttons(jd_number):
    """Return Telegram inline_keyboard rows: Save / Hide / Apply."""
    j = str(jd_number).strip()
    return [
        [
            {"text": "⭐ Save", "callback_data": f"act:save:{j}"},
            {"text": "🚫 Hide", "callback_data": f"act:hide:{j}"},
            {"text": "✅ Applied", "callback_data": f"act:apply:{j}"},
        ]
    ]


def telegram_send_card(token, chat_id, row, source, include_actions=False):
    if not token or not chat_id:
        return False
    text = format_telegram_card(row, source)
    api_url = f"{TELEGRAM_API}/bot{token}/sendMessage"
    payload_d = {
        "chat_id": str(chat_id), "text": text, "parse_mode": "HTML",
        "disable_web_page_preview": True,
    }
    if include_actions:
        jd = row.get("JD Number") or ""
        if jd:
            payload_d["reply_markup"] = {
                "inline_keyboard": build_job_action_buttons(jd)
            }
    payload = json.dumps(payload_d).encode("utf-8")
    req = urllib.request.Request(
        api_url, data=payload, method="POST",
        headers={"Content-Type": "application/json"},
    )
    try:
        with urllib.request.urlopen(req, timeout=20) as resp:
            return resp.status == 200
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", errors="replace")[:300]
        if e.code == 429:
            try:
                wait = int(
                    json.loads(body).get("parameters", {}).get("retry_after", 5)
                )
            except Exception:
                wait = 5
            print(f"    telegram rate-limited, sleeping {wait}s")
            time.sleep(wait + 0.5)
        else:
            print(f"    telegram HTTP {e.code}: {body}")
        return False
    except Exception as e:
        print(f"    telegram send failed: {e}")
        return False


def telegram_test_ping(token, chat_id):
    if not token or not chat_id:
        return False, "Missing token or chat ID"
    text = (
        "<b>✅ Job Scraper test message</b>\n"
        "Your bot token + chat ID are correctly configured.\n"
        f"<i>{datetime.now().isoformat(timespec='seconds')}</i>"
    )
    api_url = f"{TELEGRAM_API}/bot{token}/sendMessage"
    payload = json.dumps({
        "chat_id": str(chat_id), "text": text, "parse_mode": "HTML",
    }).encode("utf-8")
    req = urllib.request.Request(
        api_url, data=payload, method="POST",
        headers={"Content-Type": "application/json"},
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            return resp.status == 200, "OK"
    except urllib.error.HTTPError as e:
        return False, f"HTTP {e.code}: {e.read().decode('utf-8', errors='replace')[:200]}"
    except Exception as e:
        return False, str(e)


def _load_cv_profile(args):
    """Load CV from args.cv if provided; print summary; return profile or None."""
    path = (getattr(args, "cv", "") or "").strip()
    if not path:
        return None
    if cv_match is None:
        print("  WARN: cv_match module not available; skipping CV scoring.")
        return None
    profile = cv_match.load_cv(path)
    if profile is None:
        print(f"  WARN: could not load CV from {path}; skipping CV scoring.")
        return None
    print(f"  {profile.summary()}")
    return profile


def _open_csv(args, source):
    """Open a per-run CSV writer if --csv enabled; return (writer, file, path)."""
    if not getattr(args, "csv", True):
        return None, None, None
    path = Path(
        getattr(args, "output", None)
        or f"{source}_{args.keyword.replace(' ', '_')}_"
        f"{datetime.now():%Y%m%d_%H%M%S}.csv"
    )
    f = open(path, "w", encoding="utf-8-sig", newline="")
    w = csv.DictWriter(f, fieldnames=FIELDNAMES)
    w.writeheader()
    return w, f, path


def _finalize_scrape(args, csv_path, db, tg_state, dup_count, stopped, source):
    prefix = "Stopped by user." if stopped else "Done."
    below = tg_state.get("count_below_threshold", 0)
    sent = tg_state["count_sent"]
    summary_parts = [
        f"new={tg_state['count_new']}",
        f"dup_skipped={dup_count}",
        f"telegram_sent={sent}",
    ]
    if below:
        summary_parts.append(f"below_threshold={below}")
    tg_enabled = bool(getattr(args, "telegram_enabled", False))
    if not tg_enabled and tg_state['count_new'] > 0:
        summary_parts.append("(telegram disabled)")
    elif tg_enabled and tg_state['count_new'] > 0 and sent == 0 and below == 0:
        summary_parts.append("(check token/chat_id)")
    summary = ", ".join(summary_parts)
    pieces = []
    if csv_path:
        pieces.append(f"CSV: {csv_path.resolve()}")
    if db.path:
        pieces.append(f"Master: {db.path}")
    out = " | ".join(pieces) if pieces else "(no output)"
    print(f"\n{prefix} {summary}. {out}")
    return csv_path or db.path


def apply_cv_score(row, cv_profile):
    """Compute Match Score + Match Keywords for one row, in-place."""
    if cv_profile is None or cv_match is None:
        return
    score, matched = cv_match.score_job(cv_profile, row)
    row["Match Score"] = score
    row["Match Keywords"] = ", ".join(matched)


def process_new_row(row, args, source, db, tg_state):
    """Append a new row to master + send Telegram card if enabled.

    Returns True if persisted as new, False if already known (skipped).
    """
    jid = str(row.get("JD Number", "")).strip()
    if db.path and db.has(jid):
        return False
    db.append(row, source)
    tg_state["count_new"] += 1
    if tg_state["count_new"] % 5 == 0:
        db.save()

    if getattr(args, "telegram_enabled", False):
        token = getattr(args, "telegram_token", "") or ""
        chat = getattr(args, "telegram_chat_id", "") or ""
        if token and chat:
            # CV match-score threshold gating
            threshold = float(getattr(args, "match_threshold", 0) or 0)
            score = row.get("Match Score")
            below = (
                threshold > 0
                and isinstance(score, (int, float))
                and score < threshold
            )
            limit = int(getattr(args, "telegram_max", 0) or 0)
            if below:
                tg_state["count_below_threshold"] = (
                    tg_state.get("count_below_threshold", 0) + 1
                )
            elif limit == 0 or tg_state["count_sent"] < limit:
                include_actions = bool(getattr(args, "include_actions", False))
                if telegram_send_card(
                    token, chat, row, source, include_actions=include_actions
                ):
                    tg_state["count_sent"] += 1
                time.sleep(float(getattr(args, "telegram_delay", 1.5)))
    return True


def fetch_redux(session, url):
    """Fetch a JobsDB page and return the parsed window.SEEK_REDUX_DATA object."""
    r = session.get(url, headers=HEADERS, timeout=30)
    if r.status_code != 200:
        raise RuntimeError(f"HTTP {r.status_code} for {url}")

    m = re.search(
        r'<script[^>]*data-automation="server-state"[^>]*>(.*?)</script>',
        r.text,
        re.DOTALL,
    )
    if not m:
        raise RuntimeError(
            f"server-state script not found at {url}. "
            "JobsDB site structure may have changed."
        )
    body = m.group(1)

    m2 = re.search(r'window\.SEEK_REDUX_DATA\s*=\s*', body)
    if not m2:
        raise RuntimeError("SEEK_REDUX_DATA not found in server-state script")
    start = m2.end()

    depth = 0
    in_str = False
    esc = False
    end = None
    for i in range(start, len(body)):
        ch = body[i]
        if in_str:
            if esc:
                esc = False
            elif ch == "\\":
                esc = True
            elif ch == '"':
                in_str = False
        else:
            if ch == '"':
                in_str = True
            elif ch == '{':
                depth += 1
            elif ch == '}':
                depth -= 1
                if depth == 0:
                    end = i + 1
                    break
    if end is None:
        raise RuntimeError("Could not brace-match SEEK_REDUX_DATA object")
    return json.loads(body[start:end])


def strip_html(s):
    """Convert JobsDB-style HTML JD into clean plain text."""
    if not s:
        return ""
    s = s.replace("\r", "")
    s = re.sub(r"<br\s*/?>", "\n", s, flags=re.I)
    s = re.sub(r"</(p|div|h[1-6])\s*>", "\n", s, flags=re.I)
    s = re.sub(r"<(p|div|h[1-6])(\s[^>]*)?>", "\n", s, flags=re.I)
    s = re.sub(r"<li(\s[^>]*)?>", "\n- ", s, flags=re.I)
    s = re.sub(r"<[^>]+>", "", s)
    s = html_mod.unescape(s)
    s = re.sub(r"[ \t]+", " ", s)
    s = re.sub(r"\n[ \t]+", "\n", s)
    s = re.sub(r"\n{3,}", "\n\n", s)
    return s.strip()


def is_heading(line, pat):
    """A line is a section heading if short, keyword-matching, and either
    colon-ended or only a few words."""
    s = line.strip()
    if not s or len(s) > 60:
        return False
    has_colon = s.endswith(":") or s.endswith("：")
    t = s.rstrip(":：").strip()
    t = re.sub(r"^[-•*]\s*", "", t)
    if not has_colon and len(t.split()) > 5:
        return False
    return bool(pat.search(t))


def split_jd_sections(text):
    """Split cleaned JD text into (responsibilities, requirements, benefits, apply)."""
    if not text:
        return "", "", "", ""
    lines = text.split("\n")
    buckets = {
        "resp": [],
        "req": [],
        "benefits": [],
        "apply": [],
        "_pre": [],
    }
    current = "_pre"

    for line in lines:
        if is_heading(line, APPLY_PATTERN):
            current = "apply"
            continue
        if is_heading(line, BENEFITS_PATTERN):
            current = "benefits"
            continue
        if is_heading(line, RESP_PATTERN):
            current = "resp"
            continue
        if is_heading(line, REQ_PATTERN):
            current = "req"
            continue
        buckets[current].append(line)

    resp = "\n".join(buckets["resp"]).strip()
    req = "\n".join(buckets["req"]).strip()
    benefits = "\n".join(buckets["benefits"]).strip()
    apply_text = "\n".join(buckets["apply"]).strip()

    if not benefits:
        benefits = fallback_benefits(text)
    if not apply_text:
        apply_text = fallback_apply(text)
    return resp, req, benefits, apply_text


def fallback_benefits(text):
    """Scan JD for inline benefits sentences when no heading was matched.
    Prefer benefits-only content; if mixed with apply text, trim the apply tail."""
    if not text:
        return ""
    pat = re.compile(
        r"(we\s+(offer|provide)|5[- ]?day\s+work|annual\s+leave|medical\s+(plan|insurance|coverage)|"
        r"discretionary\s+bonus|year[- ]end\s+bonus|performance\s+bonus|"
        r"competitive\s+(salary|remuneration|package)|attractive\s+(salary|remuneration|package)|"
        r"福\s*利|待\s*遇|薪酬\s*(福利|待遇)|花紅|年終|醫療|有薪假|公眾假)",
        re.IGNORECASE,
    )
    paras = [p.strip() for p in re.split(r"\n\s*\n", text) if p.strip()]

    pure = [p for p in paras if pat.search(p) and not APPLY_PATTERN.search(p)]
    if pure:
        return "\n\n".join(pure[:3])

    out = []
    for p in paras:
        if not pat.search(p):
            continue
        m = APPLY_PATTERN.search(p)
        out.append(p[: m.start()].rstrip(" .,;") if m else p)
        if len(out) >= 2:
            break
    if out:
        return "\n\n".join(s for s in out if s)

    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    line_hits = [
        ln for ln in lines
        if pat.search(ln) and not APPLY_PATTERN.search(ln)
    ]
    return "\n".join(line_hits[:8])


def fallback_apply(text):
    """If no 'How to apply' heading found, scan the tail of the JD for
    application-related sentences (emails, 'interested parties...', etc)."""
    if not text:
        return ""
    paras = [p.strip() for p in re.split(r"\n\s*\n", text) if p.strip()]
    found = []
    for p in paras[-6:]:
        if APPLY_PATTERN.search(p) or EMAIL_RE.search(p):
            found.append(p)
    if found:
        return "\n\n".join(found)

    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    hits = [ln for ln in lines[-15:] if APPLY_PATTERN.search(ln) or EMAIL_RE.search(ln)]
    return "\n".join(hits)


def extract_summary_row(job):
    job_id = str(job.get("id") or "")

    company = job.get("companyName") or ""
    if not company:
        adv = job.get("advertiser") or {}
        if isinstance(adv, dict):
            company = adv.get("description") or ""

    locations = job.get("locations") or []
    location = ""
    if locations and isinstance(locations[0], dict):
        location = locations[0].get("label") or ""

    classifications = job.get("classifications") or []
    classification = ""
    if classifications and isinstance(classifications[0], dict):
        c = classifications[0]
        cls = (c.get("classification") or {}).get("description", "")
        sub = (c.get("subclassification") or {}).get("description", "")
        classification = f"{cls} / {sub}" if cls and sub else (cls or sub)

    work_types = job.get("workTypes") or []
    work_type = ", ".join(w for w in work_types if w)

    return {
        "JD Number": job_id,
        "Job Title": job.get("title") or "",
        "Company": company,
        "Salary": job.get("salaryLabel") or "",
        "Location": location,
        "Posted Date": job.get("listingDate") or "",
        "Posted (display)": job.get("listingDateDisplay") or "",
        "Classification": classification,
        "Work Type": work_type,
        "Responsibilities": "",
        "Requirements": "",
        "Benefits": "",
        "How to apply": "",
        "URL": f"{BASE}/job/{job_id}" if job_id else "",
    }


def fetch_full_jd(session, job_id):
    """Fetch the job detail page and return (responsibilities, requirements,
    benefits, how_to_apply)."""
    url = f"{BASE}/job/{job_id}"
    try:
        data = fetch_redux(session, url)
    except Exception as e:
        print(f"    detail fetch failed for {job_id}: {e}")
        return "", "", "", ""

    job = (
        ((data.get("jobdetails") or {}).get("result") or {}).get("job") or {}
    )
    raw = job.get("content") or job.get("content2") or job.get("abstract")
    if not raw:
        return "", "", "", ""
    full = strip_html(raw)
    return split_jd_sections(full)


def parse_at(s):
    s = s.strip()
    now = datetime.now()
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    for fmt in ("%H:%M", "%H:%M:%S"):
        try:
            t = datetime.strptime(s, fmt).time()
            target = now.replace(
                hour=t.hour, minute=t.minute, second=t.second, microsecond=0
            )
            if target <= now:
                target += timedelta(days=1)
            return target
        except ValueError:
            pass
    raise argparse.ArgumentTypeError(
        f"Cannot parse --at value: {s!r}. "
        "Use 'HH:MM' or 'YYYY-MM-DD HH:MM'."
    )


def wait_until(target, stop_event=None):
    while True:
        if stop_event is not None and stop_event.is_set():
            print()
            return
        remaining = (target - datetime.now()).total_seconds()
        if remaining <= 0:
            print()
            return
        mins, secs = divmod(int(remaining), 60)
        hrs, mins = divmod(mins, 60)
        print(
            f"  waiting until {target:%Y-%m-%d %H:%M:%S} "
            f"({hrs:02d}:{mins:02d}:{secs:02d} left)",
            end="\r",
        )
        time.sleep(min(remaining, 5))


# (legacy duplicate Master/Telegram/JobSink helpers removed; see top-of-file
# MasterDB / format_telegram_card / telegram_send_card / process_new_row)


def __legacy_block_kept_for_diff(path):
    """Placeholder to anchor the legacy-removed comment; do not call."""
    p = Path(path)
    if not p.exists() or load_workbook is None:
        return set()
    try:
        wb = load_workbook(p, read_only=True, data_only=True)
    except Exception as e:
        print(f"  [!] could not open master {p}: {e}")
        return set()
    ws = wb.active
    if ws.max_row is None or ws.max_row < 2:
        wb.close()
        return set()

    headers = [c.value for c in ws[1]]
    try:
        jd_col = headers.index("JD Number")
    except ValueError:
        wb.close()
        return set()

    ids = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        v = row[jd_col] if jd_col < len(row) else None
        if v is not None:
            ids.add(str(v))
    wb.close()
    return ids


def append_to_master(path, row, source):
    """Append a single row to the master xlsx, creating it if absent."""
    if Workbook is None:
        return
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    record = {
        "Source": source,
        "Scraped At": datetime.now().isoformat(timespec="seconds"),
        **{k: row.get(k, "") for k in FIELDNAMES},
    }
    if p.exists():
        wb = load_workbook(p)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Jobs"
        ws.append(FIELDNAMES_MASTER)
        ws.freeze_panes = "A2"
    ws.append([record.get(h, "") for h in FIELDNAMES_MASTER])
    wb.save(p)
    wb.close()


# ============================================================
# Telegram
# ============================================================

EMOJI_MAP = {
    "title": "\U0001F4CB",   # 📋
    "company": "\U0001F3E2",  # 🏢
    "salary": "\U0001F4B0",   # 💰
    "location": "\U0001F4CD", # 📍
    "date": "\U0001F4C5",     # 📅
    "type": "\U0001F4BC",     # 💼
    "resp": "\U0001F4DD",     # 📝
    "link": "\U0001F517",     # 🔗
    "tag": "\U0001F310",      # 🌐
}


def _tg_escape(s):
    if s is None:
        return ""
    return html_mod.escape(str(s))[:1000]


def _unused_telegram_send_card(session, token, chat_id, row, source):
    """Legacy curl_cffi-based sender, no longer used; kept for reference."""
    if not (token and chat_id):
        return False, "missing token/chat_id"

    title = row.get("Job Title") or "(untitled)"
    parts = [f"<b>{EMOJI_MAP['title']} {_tg_escape(title)}</b>"]

    company = row.get("Company")
    if company:
        parts.append(f"{EMOJI_MAP['company']} {_tg_escape(company)}")
    salary = row.get("Salary")
    if salary:
        parts.append(f"{EMOJI_MAP['salary']} {_tg_escape(salary)}")
    location = row.get("Location")
    if location:
        parts.append(f"{EMOJI_MAP['location']} {_tg_escape(location)}")
    posted = row.get("Posted Date") or row.get("Posted (display)")
    if posted:
        parts.append(f"{EMOJI_MAP['date']} {_tg_escape(posted)}")
    work_type = row.get("Work Type")
    if work_type:
        parts.append(f"{EMOJI_MAP['type']} {_tg_escape(work_type)}")
    parts.append(f"{EMOJI_MAP['tag']} <i>{_tg_escape(source)}</i>")

    resp = (row.get("Responsibilities") or "").strip()
    if resp:
        snippet = resp.replace("\n", " ").strip()
        if len(snippet) > 240:
            snippet = snippet[:240] + "..."
        parts.append(f"\n{EMOJI_MAP['resp']} {_tg_escape(snippet)}")

    url = row.get("URL")
    if url:
        parts.append(f'\n{EMOJI_MAP["link"]} <a href="{_tg_escape(url)}">View Job</a>')

    text = "\n".join(parts)
    if len(text) > 4000:
        text = text[:3990] + "...</a>"

    api_url = f"{TELEGRAM_API}/bot{token}/sendMessage"
    try:
        r = session.post(
            api_url,
            json={
                "chat_id": chat_id,
                "text": text,
                "parse_mode": "HTML",
                "disable_web_page_preview": False,
            },
            timeout=20,
        )
    except Exception as e:
        return False, str(e)
    if r.status_code != 200:
        return False, f"HTTP {r.status_code}: {r.text[:200]}"
    try:
        body = r.json()
    except Exception:
        return False, f"non-JSON: {r.text[:200]}"
    if not body.get("ok"):
        return False, body.get("description", "unknown error")
    return True, None


# ============================================================
# JobSink -- handles dedupe + master write + telegram
# ============================================================

class JobSink:
    """Coordinates per-job side effects: dedup, master xlsx, telegram, CSV."""

    def __init__(self, args, session=None):
        self.args = args
        self.session = session
        self.master_path = getattr(args, "master", None) or None
        if self.master_path:
            self.master_path = Path(self.master_path)
        self.tg_token = (getattr(args, "telegram_token", "") or "").strip()
        self.tg_chat = (getattr(args, "telegram_chat", "") or "").strip()
        self.tg_enabled = bool(
            getattr(args, "telegram", False) and self.tg_token and self.tg_chat
        )
        self.tg_max = int(getattr(args, "max_telegram", 0) or 0)
        self.tg_delay = float(getattr(args, "telegram_delay", 1.2) or 1.2)

        self.existing_ids = set()
        self.csv_writer = None
        self.csv_file = None
        self.tg_sent = 0
        self.tg_failed = 0
        self.new_count = 0
        self.dup_count = 0

        if self.master_path:
            self.existing_ids = load_existing_jd_numbers(self.master_path)
            print(
                f"Master: {self.master_path.name} "
                f"({len(self.existing_ids)} existing IDs known)"
            )

    def open_csv(self, path):
        self.csv_file = open(path, "w", encoding="utf-8-sig", newline="")
        self.csv_writer = csv.DictWriter(self.csv_file, fieldnames=FIELDNAMES)
        self.csv_writer.writeheader()

    def close(self):
        if self.csv_file:
            self.csv_file.close()
            self.csv_file = None

    def add(self, row, source):
        """Add one row. Returns True if accepted as new, False if duplicate."""
        jd = str(row.get("JD Number") or "").strip()
        if not jd:
            return False
        if jd in self.existing_ids:
            self.dup_count += 1
            return False

        if self.csv_writer:
            self.csv_writer.writerow(row)
            if self.csv_file:
                self.csv_file.flush()

        if self.master_path:
            try:
                append_to_master(self.master_path, row, source)
            except PermissionError:
                print(
                    f"    [!] could not write to {self.master_path.name} "
                    f"-- is it open in Excel? Skipping master append."
                )
            except Exception as e:
                print(f"    [!] master append failed: {e}")

        if self.tg_enabled and (not self.tg_max or self.tg_sent < self.tg_max):
            ok, err = telegram_send_card(
                self.session, self.tg_token, self.tg_chat, row, source
            )
            if ok:
                self.tg_sent += 1
            else:
                self.tg_failed += 1
                print(f"    [TG] send failed: {err}")
            time.sleep(self.tg_delay)

        self.existing_ids.add(jd)
        self.new_count += 1
        return True

    def summary(self):
        return (
            f"new={self.new_count}, duplicates_skipped={self.dup_count}, "
            f"telegram_sent={self.tg_sent}, telegram_failed={self.tg_failed}"
        )


def parse_args():
    p = argparse.ArgumentParser(description="Scrape HK job sites by keyword.")
    p.add_argument(
        "--source", choices=SOURCES, default="jobsdb",
        help="Which website to scrape (default: jobsdb)",
    )
    p.add_argument("--keyword", default="Accountant")
    p.add_argument("--location", default="",
                   help='Location filter, e.g. "Central and Western District", '
                        '"Tsim Sha Tsui", or "Aberdeen" for CTgoodjobs.')
    p.add_argument("--max-pages", type=int, default=0)
    p.add_argument("--full-jd", action="store_true", default=True)
    p.add_argument("--no-full-jd", dest="full_jd", action="store_false")
    p.add_argument("--delay", type=float, default=1.5)
    p.add_argument("--output", default=None,
                   help="Per-run CSV path (default: auto-named)")
    p.add_argument("--no-csv", dest="csv", action="store_false",
                   default=True, help="Skip the per-run CSV output")
    p.add_argument("--master", default=DEFAULT_MASTER_FILENAME,
                   help="Master Excel file (default: jobs_master.xlsx, "
                        "use empty string to disable)")
    p.add_argument("--telegram", dest="telegram_enabled",
                   action="store_true", default=False,
                   help="Send each new job to Telegram as it is scraped")
    p.add_argument("--telegram-token", default="",
                   help="Telegram Bot Token from BotFather")
    p.add_argument("--telegram-chat-id", default="",
                   help="Telegram Chat ID (numeric)")
    p.add_argument("--telegram-max", type=int, default=0,
                   help="Cap total Telegram messages per run (0 = unlimited)")
    p.add_argument("--telegram-delay", type=float, default=1.5,
                   help="Seconds between Telegram messages (default 1.5)")
    p.add_argument("--cv", default="",
                   help="Path to your CV (PDF or TXT) for match scoring")
    p.add_argument("--include-actions", action="store_true", default=False,
                   help="Add Save / Hide / Apply inline buttons to Telegram cards "
                        "(only useful when bot_listener.py is running)")
    p.add_argument("--match-threshold", type=float, default=0.0,
                   help="Minimum match score (0-100) to push to Telegram. "
                        "Master/CSV still capture everything. Default 0.")
    p.add_argument("--at", type=parse_at, default=None)
    return p.parse_args()


def build_search_url(keyword, location, page):
    keyword_slug = keyword.strip().replace(" ", "-")
    parts = [BASE, f"{keyword_slug}-jobs"]
    if location and location.strip():
        loc_slug = re.sub(r"\s+", "-", location.strip())
        loc_slug = re.sub(r"[,]+", "", loc_slug)
        parts.append(f"in-{loc_slug}")
    return "/".join(parts) + f"?page={page}"


# ============================================================
# CTgoodjobs source
# ============================================================

def ct_slug(s):
    s = re.sub(r"[,/]+", " ", s)
    s = re.sub(r"\s+", "-", s.strip())
    return s.lower()


def ct_build_search_url(keyword, loc_path, page):
    """Build a CT search URL given a pre-resolved location path fragment.

    loc_path: None, or e.g. 'in-new-territories/tseung-kwan-o'.
    """
    kw = ct_slug(keyword) or "all"
    url = f"{CT_BASE}/jobs/{kw}-jobs"
    if loc_path:
        url += f"/{loc_path}"
    if page > 1:
        url += f"?page={page}"
    return url


def ct_resolve_location_path(session, keyword, location):
    """Find the CT URL fragment that actually returns results for this location.

    CT districts live under one of four region prefixes — we don't know which
    until we try. Returns None if nothing works.
    """
    if not location or not location.strip():
        return None
    loc = ct_slug(location).lstrip("/")
    if not loc:
        return None
    if loc.startswith("in-"):
        loc = loc[3:]

    if "/" in loc:
        candidates = [loc]
    else:
        candidates = [loc]
        for region in CT_HK_REGIONS:
            if region != loc:
                candidates.append(f"{region}/{loc}")

    kw = ct_slug(keyword) or "all"
    for c in candidates:
        url = f"{CT_BASE}/jobs/{kw}-jobs/in-{c}"
        try:
            r = session.get(url, headers=HEADERS, timeout=30, allow_redirects=False)
        except Exception:
            continue
        if r.status_code == 200:
            return f"in-{c}"
    return None


def ct_extract_ld(text, type_name):
    """Yield parsed JSON-LD blocks of the given @type from raw HTML."""
    for body in re.findall(
        r'<script[^>]*type="application/ld\+json"[^>]*>(.*?)</script>',
        text, re.DOTALL,
    ):
        try:
            data = json.loads(body)
        except Exception:
            continue
        if isinstance(data, dict) and data.get("@type") == type_name:
            yield data


def ct_parse_job_urls_from_html(text):
    urls = []
    for data in ct_extract_ld(text, "ItemList"):
        for item in data.get("itemListElement") or []:
            if isinstance(item, dict) and item.get("url"):
                urls.append(item["url"])
        break
    return urls


def ct_parse_salary(base_salary):
    if not isinstance(base_salary, dict):
        return ""
    value = base_salary.get("value") or {}
    if not isinstance(value, dict):
        return ""
    v = value.get("value")
    if not v or str(v).upper() in ("N/A", "0", ""):
        return ""
    cur = base_salary.get("currency", "")
    unit = value.get("unitText", "")
    parts = [cur, str(v)]
    if unit and unit.upper() != "N/A":
        parts.append(f"/ {unit}")
    return " ".join(p for p in parts if p)


def ct_parse_location(job_location):
    if isinstance(job_location, list) and job_location:
        first = job_location[0]
    elif isinstance(job_location, dict):
        first = job_location
    else:
        return ""
    addr = first.get("address") if isinstance(first, dict) else None
    if not isinstance(addr, dict):
        return ""
    locality = addr.get("addressLocality") or ""
    street = addr.get("streetAddress") or ""
    if locality and street and street != locality:
        return f"{street}, {locality}"
    return locality or street


def ct_parse_emp_type(emp):
    if isinstance(emp, list):
        items = [str(x) for x in emp]
    elif isinstance(emp, str):
        items = re.findall(r"[A-Z_]{3,}", emp) or [emp]
    else:
        items = []
    return ", ".join(t.replace("_", " ").title() for t in items if t)


def ct_fetch_full_job(session, job_url):
    """Fetch one CT job detail page; return a full row dict."""
    r = session.get(job_url, headers=HEADERS, timeout=30)
    if r.status_code != 200:
        raise RuntimeError(f"HTTP {r.status_code} for {job_url}")

    job_data = None
    for data in ct_extract_ld(r.text, "JobPosting"):
        job_data = data
        break
    if job_data is None:
        raise RuntimeError(f"No JobPosting JSON-LD at {job_url}")

    m = re.search(r"/job/(\d+)", job_url)
    job_id = m.group(1) if m else ""

    org = job_data.get("hiringOrganization") or {}
    company = org.get("name", "") if isinstance(org, dict) else ""

    desc_html = job_data.get("description", "") or ""
    full_text = strip_html(desc_html)
    resp, req, benefits, apply_text = split_jd_sections(full_text)

    return {
        "JD Number": job_id,
        "Job Title": job_data.get("title", ""),
        "Company": company,
        "Salary": ct_parse_salary(job_data.get("baseSalary")),
        "Location": ct_parse_location(job_data.get("jobLocation")),
        "Posted Date": job_data.get("datePosted", ""),
        "Posted (display)": "",
        "Classification": "",
        "Work Type": ct_parse_emp_type(job_data.get("employmentType")),
        "Responsibilities": resp,
        "Requirements": req,
        "Benefits": benefits,
        "How to apply": apply_text,
        "URL": job_url,
    }


def scrape_ctgoodjobs(args, stop_event=None):
    session = requests.Session(impersonate=IMPERSONATE)
    try:
        session.get(CT_BASE, headers=HEADERS, timeout=30)
    except Exception as e:
        print(f"  warm-up request failed (continuing anyway): {e}")

    location = (getattr(args, "location", "") or "").strip()
    loc_path = None
    if location:
        print(f"Resolving CT location for {location!r}...")
        loc_path = ct_resolve_location_path(session, args.keyword, location)
        if loc_path:
            print(f"  [OK] Using path /{loc_path}/")
        else:
            print(
                f"  [!] Could not resolve {location!r} on CTgoodjobs.\n"
                f"    Try a recognized district such as: Aberdeen, Central, "
                f"Tsim Sha Tsui, Mong Kok, Tseung Kwan O,\n"
                f"    Sha Tin, Tsuen Wan, Tai Po, Kowloon Bay, "
                f"Kwun Tong, Causeway Bay, etc.\n"
                f"    ABORTING -- no scraping done to avoid wrong-location results."
            )
            return None

    cv_profile = _load_cv_profile(args)

    db = MasterDB(getattr(args, "master", None) or None)
    try:
        db.open()
    except RuntimeError as e:
        print(f"  WARN: {e}")
        db.path = None

    csv_writer, csv_file, output_path = _open_csv(args, "ctgoodjobs")
    tg_state = {"count_new": 0, "count_sent": 0, "count_below_threshold": 0}
    dup_count = 0
    zero_streak = 0
    stopped = False

    try:
        page = 1
        while True:
            if stop_event is not None and stop_event.is_set():
                stopped = True
                break
            if args.max_pages and page > args.max_pages:
                break

            url = ct_build_search_url(args.keyword, loc_path, page)
            print(f"Fetching page {page}: {url}")
            try:
                r = session.get(
                    url, headers=HEADERS, timeout=30, allow_redirects=False
                )
                if r.status_code != 200:
                    print(
                        f"  HTTP {r.status_code} (redirect to "
                        f"{r.headers.get('Location', '?')}); stopping."
                    )
                    break
                job_urls = ct_parse_job_urls_from_html(r.text)
            except Exception as e:
                print(f"  ERROR: {e}")
                break

            if not job_urls:
                print("  No jobs in this page payload. Stopping.")
                break

            page_new = 0
            for jurl in job_urls:
                if stop_event is not None and stop_event.is_set():
                    stopped = True
                    break
                m = re.search(r"/job/(\d+)", jurl)
                jid = m.group(1) if m else jurl
                if db.has(jid):
                    dup_count += 1
                    continue

                if args.full_jd:
                    try:
                        row = ct_fetch_full_job(session, jurl)
                    except Exception as e:
                        print(f"    detail fetch failed for {jid}: {e}")
                        continue
                    time.sleep(args.delay)
                else:
                    row = {fn: "" for fn in FIELDNAMES}
                    row["JD Number"] = jid
                    row["URL"] = jurl

                apply_cv_score(row, cv_profile)

                if csv_writer:
                    csv_writer.writerow(row)
                    csv_file.flush()
                if process_new_row(row, args, "ctgoodjobs", db, tg_state):
                    page_new += 1

            print(
                f"  page {page}: {page_new} new "
                f"(cumulative new: {tg_state['count_new']}, "
                f"dup skipped this run: {dup_count})"
            )

            if stopped:
                break
            if page_new == 0:
                zero_streak += 1
                if not args.max_pages and zero_streak >= 3:
                    print(
                        f"  {zero_streak} consecutive pages all-duplicates; "
                        "stopping."
                    )
                    break
                print(f"  page {page}: all jobs already in master; continuing...")
            else:
                zero_streak = 0

            page += 1
            time.sleep(args.delay)
    finally:
        if csv_file:
            csv_file.close()
        db.save()

    return _finalize_scrape(
        args, output_path, db, tg_state, dup_count, stopped, "ctgoodjobs"
    )


# ============================================================
# cpjobs source
# ============================================================

CP_HEADERS = {
    "Accept": "application/json",
    "Accept-Language": "en-HK,en;q=0.9,zh-HK;q=0.8",
    "Origin": "https://www.cpjobs.com",
    "Referer": "https://www.cpjobs.com/",
}


def cp_format_salary(p):
    minp = p.get("minsalary")
    maxp = p.get("maxsalary")
    show = p.get("showsalary")
    if show == "N":
        return ""
    # cpjobs sometimes returns salaries as strings ("30000") rather than ints,
    # which crashes f"{x:,}" with "Cannot specify ',' with 's'". Coerce safely.
    def _to_int(v):
        if v is None or v == "":
            return None
        if isinstance(v, (int, float)):
            return int(v)
        try:
            return int(str(v).replace(",", "").strip())
        except (TypeError, ValueError):
            return None
    minp = _to_int(minp)
    maxp = _to_int(maxp)
    if not minp and not maxp:
        return ""
    if minp and maxp:
        return f"HKD {minp:,} - {maxp:,}"
    if minp:
        return f"HKD {minp:,}+"
    return f"up to HKD {maxp:,}"


def _cp_text(v):
    """Coerce a cpjobs description field (str / dict / list) into plain text."""
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    if isinstance(v, dict):
        for k in ("desc", "description", "name", "label", "value", "en"):
            if v.get(k):
                return str(v[k]).strip()
        return ""
    if isinstance(v, list):
        return ", ".join(filter(None, (_cp_text(i) for i in v)))
    return str(v).strip()


def cp_format_location(p):
    parts = []
    for k in ("locationdesc", "citydesc", "countrydesc"):
        s = _cp_text(p.get(k))
        if s and s not in parts:
            parts.append(s)
    return " / ".join(parts)


def cp_format_classification(p):
    parts = []
    for k in ("industrydesc", "subfielddesc"):
        s = _cp_text(p.get(k))
        if s and s not in parts:
            parts.append(s)
    return " / ".join(parts)


def cp_format_posted(p):
    pub = p.get("publisheddate") or {}
    ts = pub.get("timestamp")
    if not ts:
        return ""
    try:
        return datetime.fromtimestamp(int(ts)).isoformat()
    except Exception:
        return ""


def cp_position_to_row(p):
    job_id = str(p.get("id") or "")
    desc_html = p.get("description") or ""
    full_text = strip_html(desc_html)
    resp, req, benefits, apply_text = split_jd_sections(full_text)

    return {
        "JD Number": job_id,
        "Job Title": _cp_text(p.get("jobtitletext")),
        "Company": _cp_text(p.get("displayname")),
        "Salary": cp_format_salary(p),
        "Location": cp_format_location(p),
        "Posted Date": cp_format_posted(p),
        "Posted (display)": "",
        "Classification": cp_format_classification(p),
        "Work Type": _cp_text(p.get("employmentstatusdesc")),
        "Responsibilities": resp,
        "Requirements": req,
        "Benefits": benefits,
        "How to apply": apply_text,
        "URL": p.get("url") or f"{CP_BASE}/job/{job_id}",
    }


def cp_fetch_search_page(session, keyword, cityid, offset):
    params = {"keyword": keyword, "offset": offset}
    if cityid:
        params["cityid"] = cityid
    url = f"{CP_API}/positions"
    r = session.get(url, params=params, headers=CP_HEADERS, timeout=30)
    if r.status_code != 200:
        raise RuntimeError(f"HTTP {r.status_code} for {r.url}")
    data = r.json()
    return data.get("positions") or [], data.get("total", 0), r.url


def cp_fetch_full_position(session, position_id):
    url = f"{CP_API}/positions/{position_id}"
    r = session.get(url, headers=CP_HEADERS, timeout=30)
    if r.status_code != 200:
        raise RuntimeError(f"HTTP {r.status_code} for {url}")
    return r.json()


def scrape_cpjobs(args, stop_event=None):
    session = requests.Session(impersonate=IMPERSONATE)
    try:
        session.get("https://www.cpjobs.com/hk",
                    headers={"Accept": "text/html"}, timeout=30)
    except Exception as e:
        print(f"  warm-up request failed (continuing anyway): {e}")

    location = (getattr(args, "location", "") or "").strip()
    cityid = ""
    if location:
        cityid = CP_LOCATION_MAP.get(location.title()) or CP_LOCATION_MAP.get(location)
        if cityid:
            print(f"  [OK] Using cpjobs cityid={cityid} ({location})")
        else:
            print(
                f"  [!] cpjobs only supports 4 regions: "
                f"{', '.join(CP_LOCATIONS)}.\n"
                f"    {location!r} not recognized -- "
                f"continuing without location filter."
            )

    cv_profile = _load_cv_profile(args)

    db = MasterDB(getattr(args, "master", None) or None)
    try:
        db.open()
    except RuntimeError as e:
        print(f"  WARN: {e}")
        db.path = None

    csv_writer, csv_file, output_path = _open_csv(args, "cpjobs")
    tg_state = {"count_new": 0, "count_sent": 0, "count_below_threshold": 0}
    dup_count = 0
    zero_streak = 0
    stopped = False
    api_total = None
    PAGE_SIZE = 25

    try:
        page = 1
        while True:
            if stop_event is not None and stop_event.is_set():
                stopped = True
                break
            if args.max_pages and page > args.max_pages:
                break

            offset = (page - 1) * PAGE_SIZE
            if api_total is not None and offset >= api_total:
                print("  reached end of results.")
                break

            try:
                positions, api_total, eff_url = cp_fetch_search_page(
                    session, args.keyword, cityid, offset
                )
            except Exception as e:
                print(f"  ERROR: {e}")
                break
            print(f"Fetching page {page}: {eff_url}")
            if page == 1:
                print(f"  API reports total={api_total} matching jobs")

            if not positions:
                print("  No positions in this page payload. Stopping.")
                break

            page_new = 0
            for p in positions:
                if stop_event is not None and stop_event.is_set():
                    stopped = True
                    break
                pid = str(p.get("id") or "")
                if not pid:
                    continue
                if db.has(pid):
                    dup_count += 1
                    continue

                if args.full_jd:
                    try:
                        detail = cp_fetch_full_position(session, pid)
                    except Exception as e:
                        print(f"    detail fetch failed for {pid}: {e}")
                        detail = p
                    row = cp_position_to_row(detail)
                    time.sleep(args.delay)
                else:
                    row = cp_position_to_row(p)

                apply_cv_score(row, cv_profile)
                if csv_writer:
                    csv_writer.writerow(row)
                    csv_file.flush()
                if process_new_row(row, args, "cpjobs", db, tg_state):
                    page_new += 1

            print(
                f"  page {page}: {page_new} new "
                f"(cumulative new: {tg_state['count_new']}, "
                f"dup skipped this run: {dup_count})"
            )

            if stopped:
                break
            if page_new == 0:
                zero_streak += 1
                if not args.max_pages and zero_streak >= 3:
                    print(
                        f"  {zero_streak} consecutive pages all-duplicates; "
                        "stopping."
                    )
                    break
                print(f"  page {page}: all jobs already in master; continuing...")
            else:
                zero_streak = 0

            page += 1
            time.sleep(args.delay)
    finally:
        if csv_file:
            csv_file.close()
        db.save()

    return _finalize_scrape(
        args, output_path, db, tg_state, dup_count, stopped, "cpjobs"
    )


def scrape(args, stop_event=None):
    """Dispatch to the right source-specific scraper."""
    source = (getattr(args, "source", "") or "jobsdb").lower()
    if source == "cpjobs":
        return scrape_cpjobs(args, stop_event=stop_event)
    if source == "ctgoodjobs":
        return scrape_ctgoodjobs(args, stop_event=stop_event)
    if source == "jobsdb":
        return scrape_jobsdb(args, stop_event=stop_event)
    raise ValueError(f"Unknown source: {source!r}; expected one of {SOURCES}")


def scrape_jobsdb(args, stop_event=None):
    session = requests.Session(impersonate=IMPERSONATE)
    try:
        session.get(BASE, headers=HEADERS, timeout=30)
    except Exception as e:
        print(f"  warm-up request failed (continuing anyway): {e}")

    cv_profile = _load_cv_profile(args)

    db = MasterDB(getattr(args, "master", None) or None)
    try:
        db.open()
    except RuntimeError as e:
        print(f"  WARN: {e}")
        db.path = None

    csv_writer, csv_file, output_path = _open_csv(args, "jobsdb")
    tg_state = {"count_new": 0, "count_sent": 0, "count_below_threshold": 0}
    dup_count = 0
    zero_streak = 0
    stopped = False

    try:
        page = 1
        while True:
            if stop_event is not None and stop_event.is_set():
                stopped = True
                break
            if args.max_pages and page > args.max_pages:
                break

            url = build_search_url(
                args.keyword, getattr(args, "location", "") or "", page
            )
            print(f"Fetching page {page}: {url}")
            try:
                data = fetch_redux(session, url)
            except Exception as e:
                print(f"  ERROR: {e}")
                break

            jobs = (
                (data.get("results") or {}).get("results", {}).get("jobs") or []
            )
            if not jobs:
                print("  No jobs in this page payload. Stopping.")
                break

            page_new = 0
            for j in jobs:
                if stop_event is not None and stop_event.is_set():
                    stopped = True
                    break
                row = extract_summary_row(j)
                jid = row.get("JD Number")
                if not jid:
                    continue
                if db.has(jid):
                    dup_count += 1
                    continue

                if args.full_jd:
                    resp, req, benefits, apply_text = fetch_full_jd(session, jid)
                    row["Responsibilities"] = resp
                    row["Requirements"] = req
                    row["Benefits"] = benefits
                    row["How to apply"] = apply_text
                    time.sleep(args.delay)

                apply_cv_score(row, cv_profile)
                if csv_writer:
                    csv_writer.writerow(row)
                    csv_file.flush()
                if process_new_row(row, args, "jobsdb", db, tg_state):
                    page_new += 1

            print(
                f"  page {page}: {page_new} new "
                f"(cumulative new: {tg_state['count_new']}, "
                f"dup skipped this run: {dup_count})"
            )

            if stopped:
                break
            if page_new == 0:
                zero_streak += 1
                if not args.max_pages and zero_streak >= 3:
                    print(
                        f"  {zero_streak} consecutive pages all-duplicates; "
                        "stopping."
                    )
                    break
                print(f"  page {page}: all jobs already in master; continuing...")
            else:
                zero_streak = 0

            page += 1
            time.sleep(args.delay)
    finally:
        if csv_file:
            csv_file.close()
        db.save()

    return _finalize_scrape(
        args, output_path, db, tg_state, dup_count, stopped, "jobsdb"
    )


def main():
    args = parse_args()
    if args.at:
        print(f"Scheduled run at {args.at:%Y-%m-%d %H:%M:%S}")
        wait_until(args.at)
    scrape(args)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nInterrupted by user. Partial results saved.")
        sys.exit(1)
